# -*- coding: utf-8 -*-
"""
Created on Wed Jun 25 09:10:40 2025

@author: kccheng
"""

import openpyxl
import os
import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl.utils import get_column_letter
import re
from core.excel_connector import activate_excel_window
from core.excel_scanner import refresh_data
import time

def export_formulas_to_excel(controller):
    if not controller.view.result_tree.get_children():
        messagebox.showinfo("No Data", "There is no data to export in the list.")
        return
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
        title="Save Exported Formulas As"
    )
    if not file_path:
        return
    workbook = openpyxl.Workbook()
    sheet_name = "Formulas"
    if controller.worksheet:
        sheet_name = controller.worksheet.Name
    sheet = workbook.active
    sheet.title = sheet_name
    sheet['A1'] = "Address"
    sheet['B1'] = "Formula Content"
    sheet.column_dimensions[get_column_letter(2)].number_format = '@'
    address_idx = controller.view.tree_columns.index("address")
    formula_idx = controller.view.tree_columns.index("formula")
    for i, item_id in enumerate(controller.view.result_tree.get_children()):
        values = controller.view.result_tree.item(item_id, "values")
        if len(values) > max(address_idx, formula_idx):
            sheet.cell(row=i + 2, column=1, value=values[address_idx])
            sheet.cell(row=i + 2, column=2, value="'" + values[formula_idx])
    sheet.column_dimensions[get_column_letter(1)].width = 15
    sheet.column_dimensions[get_column_letter(2)].width = 80
    backup_sheet = workbook.copy_worksheet(sheet)
    backup_sheet.title = "backup"
    workbook.save(file_path)
    messagebox.showinfo("Export Successful", f"Data successfully exported to:\n{file_path}")
    os.startfile(file_path)


def import_and_update_formulas(controller):
    root = controller.view.winfo_toplevel()
    original_topmost = root.attributes("-topmost")
    root.attributes("-topmost", True)
    
    # Use the existing progress frame from the view
    progress_frame = controller.view.progress_frame
    progress_label = controller.view.progress_label
    progress_bar = controller.view.progress_bar
    
    progress_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5) # Make it visible
    progress_label.config(text="Preparing to import formulas...")
    progress_bar.config(value=0, maximum=100)
    root.update_idletasks()
    newline = "\n"

    if not controller.workbook or not controller.worksheet:
        messagebox.showerror("Not Connected", "Please scan a worksheet first.\nThe tool needs an active worksheet to update.")
        progress_frame.grid_forget()
        root.attributes("-topmost", original_topmost)
        return

    file_path = filedialog.askopenfilename(
        filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
        title="Select File to Import Formulas From"
    )
    if not file_path:
        progress_frame.grid_forget()
        root.attributes("-topmost", original_topmost)
        return

    progress_label.config(text=f"Reading '{os.path.basename(file_path)}'...")
    progress_bar.config(value=10)
    root.update_idletasks()
    
    calc_mode_prev = None
    calc_before_save_prev = None
    enable_events_prev = None

    if controller.xl:
        try:
            calc_mode_prev = controller.xl.Calculation
            calc_before_save_prev = controller.xl.Application.CalculateBeforeSave
            enable_events_prev = controller.xl.Application.EnableEvents
            
            controller.xl.Calculation = -4135  # xlCalculationManual
            controller.xl.Application.CalculateBeforeSave = False
            controller.xl.Application.EnableEvents = False
        except Exception as e:
            print(f"Failed to set Excel application properties during import: {e}")
            calc_mode_prev = None
            calc_before_save_prev = None
            enable_events_prev = None

    import time
    start_time = time.time()
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        updates = {}
        total_rows = sheet.max_row - 1 if sheet.max_row > 1 else 0 
        progress_label.config(text=f"Scanning formulas in '{os.path.basename(file_path)}'...")
        progress_bar.config(value=20)
        root.update_idletasks()
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, values_only=True)):
            if row and row[0] and row[1]:
                address = str(row[0]).strip()
                formula = str(row[1]).strip()
                if formula.startswith("'="):
                    formula = formula[1:]
                updates[address] = formula
            if total_rows > 0:
                current_read_progress = 20 + (i / total_rows * 30)
            else:
                current_read_progress = 50
            progress_bar.config(value=current_read_progress)
            progress_label.config(text=f"Scanning formulas ({i+1}/{total_rows})...")
            root.update_idletasks()
        if not updates:
            messagebox.showinfo("No Data", "No valid Address/Formula pairs found in the selected file.")
            progress_frame.grid_forget()
            root.attributes("-topmost", original_topmost)
            return
        progress_label.config(text=f"Found {len(updates)} formulas. Confirming update...")
        progress_bar.config(value=50)
        root.update_idletasks()
        
        confirmation = messagebox.askyesno(
            "Confirm Update",
            f"You are about to update {len(updates)} formulas in the worksheet '{controller.worksheet.Name}' in '{controller.workbook.Name}'.{newline}{newline}"
            f"This action CANNOT be undone.{newline}{newline}Do you want to proceed?"
        )
        if not confirmation:
            messagebox.showinfo("Cancelled", "Update operation was cancelled.")
            progress_frame.grid_forget()
            root.attributes("-topmost", original_topmost)
            return

        activate_excel_window(controller)
        controller.worksheet.Activate()
        updated_count = 0
        error_count = 0
        errors = []
        progress_label.config(text=f"Updating {len(updates)} formulas in Excel...")
        progress_bar.config(value=60)
        root.update_idletasks()
        total_updates = len(updates)
        for i, (address, formula) in enumerate(updates.items()):
            try:
                controller.worksheet.Range(address).Formula = formula
                updated_count += 1
            except Exception as e:
                error_count += 1
                errors.append(f" - {address}: {e}")
            if total_updates > 0:
                current_update_progress = 60 + (i / total_updates * 40)
            else:
                current_update_progress = 100
            progress_bar.config(value=current_update_progress)
            progress_label.config(text=f"Updating Excel: {i+1}/{total_updates} formulas")
            root.update_idletasks()
        end_time = time.time()
        time_taken = end_time - start_time
        progress_label.config(
            text=f"Updating Excel: {total_updates}/{total_updates} formulas (Total import time: {time_taken:.2f} seconds)"
        )
        root.update_idletasks()
        
        summary_message = f"Update Complete.{newline}{newline}Successfully updated: {updated_count}{newline}Failed to update: {error_count}"
        if errors:
            summary_message += f"{newline}{newline}Errors:{newline}" + f"{newline}".join(errors[:5])
        messagebox.showinfo("Update Summary", summary_message + f"{newline}{newline}Please re-scan the worksheet to view the changes.")
    except Exception as e:
        messagebox.showerror("Import Error", f"An error occurred during import:{newline}{e}")
    finally:
        if controller.xl:
            try:
                controller.xl.Application.EnableEvents = True if enable_events_prev is None else enable_events_prev
            except Exception:
                pass
            try:
                if calc_before_save_prev is not None:
                    controller.xl.Application.CalculateBeforeSave = calc_before_save_prev
                else:
                    controller.xl.Application.CalculateBeforeSave = True
            except Exception:
                pass
            try:
                if calc_mode_prev is not None:
                    controller.xl.Calculation = calc_mode_prev
                else:
                    controller.xl.Calculation = -4105  # xlCalculationAutomatic
            except Exception:
                pass
            try:
                controller.xl.CalculateFullRebuild()
            except Exception:
                pass

        progress_bar.config(value=0)
        progress_label.config(text="")
        progress_frame.grid_forget()
        root.update_idletasks()
        root.attributes("-topmost", original_topmost)
        refresh_data(controller, None, scan_mode='quick')
