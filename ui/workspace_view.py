import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pythoncom
import win32com.client
import win32gui
import win32con
import time
import openpyxl
import os
from datetime import datetime
import threading

class AccumulateListbox(tk.Listbox):
    def __init__(self, master, **kwargs):
        super().__init__(master, selectmode=tk.EXTENDED, **kwargs)
        self._drag_start = None
        self._drag_mode = None
        self.bind("<Button-1>", self._on_click)
        self.bind("<B1-Motion>", self._on_drag)
        self.bind("<ButtonRelease-1>", self._on_release)
        self.bind('<<ListboxSelect>>', self._on_select_event)

    def _on_click(self, event):
        self.focus_set()
        idx = self.nearest(event.y)
        if idx < 0: return "break"
        self._drag_start = idx
        if self.selection_includes(idx):
            self.selection_clear(idx)
            self._drag_mode = "unselect"
        else:
            self.selection_set(idx)
            self._drag_mode = "select"
        return "break"

    def _on_drag(self, event):
        if self._drag_start is None: return
        cur = self.nearest(event.y)
        lo = min(self._drag_start, cur)
        hi = max(self._drag_start, cur)
        if self._drag_mode == "unselect":
            for idx in range(lo, hi + 1):
                self.selection_clear(idx)
        else:
            for idx in range(lo, hi + 1):
                self.selection_set(idx)
        return "break"

    def _on_release(self, event):
        self._drag_start = None
        self._drag_mode = None
        self.event_generate("<<ListboxSelect>>")
        return "break"

    def _on_select_event(self, event):
        if hasattr(self, 'external_on_selection_change'):
            self.external_on_selection_change(event)

class Workspace:
    def __init__(self, parent_frame):
        self.parent = parent_frame
        self.root = parent_frame.winfo_toplevel()
        
        self.file_names = []
        self.file_paths = []
        self.sheet_names = []
        self.active_cells = []
        self.showing_path = False
        self.target_captions = []
        
        self.setup_ui()
        self.show_names()

    def setup_ui(self):
        title_frame = tk.Frame(self.parent)
        title_frame.pack(pady=(15, 0), padx=10, anchor="w", fill='x')

        main_label = tk.Label(title_frame, text="Current Workspace:", font=("Arial", 16, "bold"))
        main_label.pack(side=tk.LEFT)

        self.count_label = tk.Label(title_frame, text="", font=("Arial", 12))
        self.count_label.pack(side=tk.LEFT, padx=(5, 0), pady=(2,0))

        main_frame = tk.Frame(self.parent)
        main_frame.pack(pady=(5, 10), padx=10, expand=True, fill='both')

        side_btn_frame = tk.Frame(main_frame, width=150)
        side_btn_frame.pack(side=tk.RIGHT, padx=(10, 0), pady=10, fill='y')
        side_btn_frame.pack_propagate(False)
        btn_props = {'width': 20, 'height': 2, 'wraplength': 140, 'font': ("Arial", 10, "bold")}

        self.refresh_btn = tk.Button(side_btn_frame, text="Refresh", **btn_props, command=self.show_names)
        self.refresh_btn.pack(pady=5, anchor='n')

        self.show_path_btn = tk.Button(side_btn_frame, text="Show Full Path", **btn_props, command=self.toggle_path)
        self.show_path_btn.pack(pady=5, anchor='n')

        activate_btn = tk.Button(side_btn_frame, text="Activate Workbook", **btn_props, command=self.activate_selected_workbooks)
        activate_btn.pack(pady=5, anchor='n')

        minimize_btn = tk.Button(side_btn_frame, text="Minimize All Workbook", **btn_props, command=self.minimize_all_excel)
        minimize_btn.pack(pady=5, anchor='n')

        save_ws_btn = tk.Button(side_btn_frame, text="Save Workspace", **btn_props, command=self.save_workspace)
        save_ws_btn.pack(pady=5, anchor='n')

        load_ws_btn = tk.Button(side_btn_frame, text="Load Workspace", **btn_props, command=self.load_workspace)
        load_ws_btn.pack(pady=5, anchor='n')

        btn1 = tk.Button(side_btn_frame, text="Save Workbook", **btn_props, command=self.save_selected_workbooks)
        btn1.pack(pady=5, anchor='n')

        btn2 = tk.Button(side_btn_frame, text="Close Workbook Without Save", **btn_props, command=lambda: self.close_selected_workbooks(False))
        btn2.pack(pady=5, anchor='n')

        btn3 = tk.Button(side_btn_frame, text="Save and Close Workbook", **btn_props, command=lambda: self.close_selected_workbooks(True))
        btn3.pack(pady=5, anchor='n')

        listbox_frame = tk.Frame(main_frame)
        listbox_frame.pack(side=tk.LEFT, fill='both', expand=True, pady=10)

        self.listbox = AccumulateListbox(
            listbox_frame,
            font=("Consolas", 12),
            height=20,
            borderwidth=2,
            relief="groove"
        )
        self.listbox.pack(expand=True, fill='both', padx=8, pady=8)
        self.listbox.external_on_selection_change = self.on_selection_change

    def get_open_excel_files(self):
        pythoncom.CoInitialize()
        excel_files, file_paths, sheet_names, active_cells = [], [], [], []
        excel = None
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            for wb in excel.Workbooks:
                excel_files.append(wb.Name)
                file_paths.append(wb.FullName)
                try:
                    sht = wb.ActiveSheet
                    sheet_names.append(sht.Name)
                    cell_addr = sht.Application.ActiveCell.Address
                    active_cells.append(cell_addr)
                except Exception:
                    sheet_names.append("")
                    active_cells.append("")
        except Exception:
            pass
        finally:
            excel = None
            pythoncom.CoUninitialize()
        return excel_files, file_paths, sheet_names, active_cells

    def show_names(self):
        def update_gui(data):
            self.file_names, self.file_paths, self.sheet_names, self.active_cells = data
            self.count_label.config(text=f"({len(self.file_names)} files open)")
            self.listbox.delete(0, tk.END)

            if not self.file_names:
                self.listbox.insert(tk.END, "There are currently no open Excel files.")
            elif self.showing_path:
                for path in self.file_paths:
                    self.listbox.insert(tk.END, path)
            else:
                for name in self.file_names:
                    self.listbox.insert(tk.END, name)
            
            self.refresh_btn.config(state=tk.NORMAL)

        def scan_in_thread():
            scan_data = self.get_open_excel_files()
            self.root.after(0, lambda: update_gui(scan_data))

        self.refresh_btn.config(state=tk.DISABLED)
        threading.Thread(target=scan_in_thread, daemon=True).start()

    def toggle_path(self):
        self.showing_path = not self.showing_path
        self.listbox.delete(0, tk.END)

        if self.showing_path:
            if self.file_paths:
                for path in self.file_paths:
                    self.listbox.insert(tk.END, path)
            else:
                self.listbox.insert(tk.END, "There are currently no open Excel files.")
            self.show_path_btn.config(text="Hide Full Path")
        else:
            if self.file_names:
                for name in self.file_names:
                    self.listbox.insert(tk.END, name)
            else:
                self.listbox.insert(tk.END, "There are currently no open Excel files.")
            self.show_path_btn.config(text="Show Full Path")
            
    def get_selected_workbooks(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices:
            messagebox.showinfo("Notice", "Please kindly select one or more Excel files before proceeding.")
            return []
        selected_workbooks = []
        for idx in selected_indices:
            if 0 <= idx < len(self.file_names):
                selected_workbooks.append((self.file_names[idx], self.file_paths[idx], self.sheet_names[idx], self.active_cells[idx]))
        return selected_workbooks

    def on_selection_change(self, event):
        selected_indices = self.listbox.curselection()
        self.target_captions = [self.file_names[idx] for idx in selected_indices if 0 <= idx < len(self.file_names)]

    def activate_selected_workbooks(self):
        if not self.target_captions:
            messagebox.showinfo("Notice", "Please kindly select one or more Excel files to activate.")
            return

        offset_x = 40
        offset_y = 40
        start_x = 100
        start_y = 100
        activated_hwnds = set()
        window_index = 0

        def enum_handler(hwnd, ctx):
            nonlocal window_index
            if win32gui.IsWindowVisible(hwnd):
                title = win32gui.GetWindowText(hwnd)
                for caption in ctx["captions"]:
                    if caption in title and hwnd not in activated_hwnds:
                        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                        win32gui.SetForegroundWindow(hwnd)
                        x = start_x + window_index * offset_x
                        y = start_y + window_index * offset_y
                        rect = win32gui.GetWindowRect(hwnd)
                        w = rect[2] - rect[0]
                        h = rect[3] - rect[1]
                        win32gui.SetWindowPos(hwnd, None, x, y, w, h, win32con.SWP_NOZORDER | win32con.SWP_SHOWWINDOW)
                        activated_hwnds.add(hwnd)
                        window_index += 1
                        time.sleep(0.05)

        ctx = {"captions": self.target_captions}
        win32gui.EnumWindows(enum_handler, ctx)

    def save_selected_workbooks(self):
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            selected = self.get_selected_workbooks()
            if not selected: return
            for name, path, _, _ in selected:
                for wb in excel.Workbooks:
                    if wb.Name == name and wb.FullName == path: wb.Save()
            messagebox.showinfo("Complete", "Your selected Excel files have been saved successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Sorry, an error occurred while saving the files:\n{str(e)}")
        finally:
            pythoncom.CoUninitialize()
            self.show_names()

    def close_selected_workbooks(self, save_before_close=False):
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
            selected = self.get_selected_workbooks()
            if not selected: return
            for name, path, _, _ in selected:
                for wb in excel.Workbooks:
                    if wb.Name == name and wb.FullName == path: wb.Close(SaveChanges=save_before_close)
            if save_before_close:
                messagebox.showinfo("Complete", "Your selected Excel files have been saved and closed successfully.")
            else:
                messagebox.showinfo("Complete", "Your selected Excel files have been closed without saving.")
        except Exception as e:
            messagebox.showerror("Error", f"Sorry, an error occurred while closing the files:\n{str(e)}")
        finally:
            pythoncom.CoUninitialize()
            self.show_names()

    def minimize_all_excel(self):
        def enum_handler(hwnd, ctx):
            if win32gui.IsWindowVisible(hwnd):
                if " - Excel" in win32gui.GetWindowText(hwnd):
                    win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
        win32gui.EnumWindows(enum_handler, None)

    def save_workspace(self):
        selected = self.get_selected_workbooks()
        if not selected:
            messagebox.showinfo("Notice", "Please kindly select one or more Excel files before saving a workspace.")
            return
        file_path = filedialog.asksaveasfilename(
            title="Save Workspace",
            defaultextension=".xlsx",
            filetypes=[("Excel Workspace", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path: return
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        base, ext = os.path.splitext(file_path)
        file_path_with_ts = f"{base}_{timestamp}{ext}"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workspace"
        ws.append(["File Path", "Sheet Name", "Cell Address"])
        for _, path, sheet, cell in selected:
            ws.append([path, sheet, cell])
        wb.save(file_path_with_ts)
        messagebox.showinfo("Success", f"The workspace has been saved successfully at:\n{file_path_with_ts}")

    def load_workspace(self):
        current_files, _, _, _ = self.get_open_excel_files()
        if current_files:
            messagebox.showwarning("Warning", "To proceed, please kindly close all currently open Excel files before loading a workspace.")
            self.show_names()
            return
        file_path = filedialog.askopenfilename(
            title="Load Workspace",
            filetypes=[("Excel Workspace", "*.xlsx"), ("All Files", "*.*")]
        )
        if not file_path or not os.path.exists(file_path): return

        def thread_job():
            pythoncom.CoInitialize()
            try:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                valid_rows = [r for r in rows if r and r[0] and os.path.exists(r[0])]
                if not valid_rows:
                    self.root.after(0, lambda: messagebox.showwarning("Warning", "Unfortunately, there are no valid file paths found in the selected workspace."))
                    self.root.after(0, self.show_names)
                    return

                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                excel.AskToUpdateLinks = False
                for r in valid_rows:
                    path, sheet, cell = (r[0], r[1] if len(r) > 1 else None, r[2] if len(r) > 2 else None)
                    try:
                        wb_xl = excel.Workbooks.Open(Filename=path, UpdateLinks=0)
                        if sheet:
                            try:
                                sht = wb_xl.Sheets(sheet)
                                sht.Activate()
                                if cell: sht.Range(cell).Select()
                            except Exception: pass
                    except Exception: pass
                excel.AskToUpdateLinks = True
                self.root.after(0, lambda: messagebox.showinfo("Complete", f"{len(valid_rows)} file(s) from the workspace have been opened for you."))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Sorry, an error occurred while loading the workspace:\n{str(e)}"))
            finally:
                pythoncom.CoUninitialize()
                self.root.after(200, self.show_names)

        threading.Thread(target=thread_job).start()