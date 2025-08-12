# -*- coding: utf-8 -*-
"""
Safe Cache System - 獨立的快取模組
分離快取邏輯，避免單一檔案過大
"""

import os
import time
import threading
import gc
from collections import OrderedDict
from openpyxl import load_workbook


class SafeWorkbookCache:
    """
    安全的工作簿快取系統
    - 強制唯讀模式防止檔案鎖定
    - 改進的記憶體管理
    - 更好的錯誤處理
    """
    
    def __init__(self, max_size=10, max_age_seconds=300):
        self.max_size = max_size
        self.max_age_seconds = max_age_seconds
        self.cache = OrderedDict()
        self.lock = threading.RLock()
        self._stats = {
            'hits': 0,
            'misses': 0,
            'evictions': 0,
            'errors': 0,
            'memory_cleanups': 0
        }
    
    def get_workbook(self, file_path, data_only=True):
        """
        安全獲取工作簿（強制唯讀模式）
        
        Args:
            file_path: Excel檔案路徑
            data_only: 是否只讀取計算值
            
        Returns:
            openpyxl.Workbook: 工作簿物件
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        normalized_path = os.path.normpath(os.path.abspath(file_path))
        cache_key = f"{normalized_path}|{data_only}"
        
        with self.lock:
            # 檢查快取
            if cache_key in self.cache:
                cached_item = self.cache[cache_key]
                
                if self._is_cache_valid(cached_item, normalized_path):
                    self.cache.move_to_end(cache_key)
                    self._stats['hits'] += 1
                    return cached_item['workbook']
                else:
                    self._safe_remove_cache_entry(cache_key)
            
            # 載入新工作簿
            return self._load_and_cache_workbook(normalized_path, cache_key, data_only)
    
    def _load_and_cache_workbook(self, file_path, cache_key, data_only):
        """安全載入並快取工作簿"""
        self._stats['misses'] += 1
        
        try:
            from openpyxl import load_workbook
            
            # 強制安全參數，但保留外部連結資訊
            workbook = load_workbook(
                filename=file_path,
                read_only=True,  # 強制唯讀
                data_only=data_only,
                keep_vba=False,
                keep_links=True  # 必須保留外部連結資訊！
            )
            
            # 創建快取項目
            cache_entry = {
                'workbook': workbook,
                'file_path': file_path,
                'file_mtime': os.path.getmtime(file_path),
                'cache_time': time.time(),
                'data_only': data_only
            }
            
            self.cache[cache_key] = cache_entry
            self._enforce_cache_limit()
            
            return workbook
            
        except Exception as e:
            self._stats['errors'] += 1
            raise Exception(f"Failed to load workbook {os.path.basename(file_path)}: {str(e)}")
    
    def _is_cache_valid(self, cached_item, file_path):
        """檢查快取有效性"""
        try:
            # 檢查年齡
            if time.time() - cached_item['cache_time'] > self.max_age_seconds:
                return False
            
            # 檢查檔案修改時間
            current_mtime = os.path.getmtime(file_path)
            if current_mtime != cached_item['file_mtime']:
                return False
            
            return True
            
        except (OSError, KeyError):
            return False
    
    def _safe_remove_cache_entry(self, cache_key):
        """安全移除快取項目"""
        try:
            if cache_key in self.cache:
                cached_item = self.cache[cache_key]
                self._cleanup_workbook(cached_item['workbook'])
                del self.cache[cache_key]
        except Exception as e:
            print(f"Warning: Error removing cache entry: {e}")
    
    def _cleanup_workbook(self, workbook):
        """安全清理工作簿"""
        try:
            if hasattr(workbook, 'close'):
                workbook.close()
            del workbook
            self._stats['memory_cleanups'] += 1
        except Exception:
            pass
    
    def _enforce_cache_limit(self):
        """執行快取大小限制"""
        while len(self.cache) > self.max_size:
            oldest_key, oldest_item = self.cache.popitem(last=False)
            self._cleanup_workbook(oldest_item['workbook'])
            self._stats['evictions'] += 1
    
    def clear(self):
        """清空快取"""
        with self.lock:
            for cache_entry in self.cache.values():
                self._cleanup_workbook(cache_entry['workbook'])
            self.cache.clear()
            # 強制垃圾回收
            gc.collect()
    
    def get_stats(self):
        """獲取統計信息"""
        with self.lock:
            total_requests = self._stats['hits'] + self._stats['misses']
            hit_rate = (self._stats['hits'] / total_requests * 100) if total_requests > 0 else 0
            
            return {
                'cache_size': len(self.cache),
                'max_size': self.max_size,
                'hit_rate_percent': round(hit_rate, 2),
                'stats': self._stats.copy(),
                'cached_files': [os.path.basename(entry['file_path']) for entry in self.cache.values()]
            }


# 全域安全快取實例
_safe_global_cache = None
_safe_cache_lock = threading.Lock()


def get_safe_global_cache():
    """獲取全域安全快取實例"""
    global _safe_global_cache
    
    if _safe_global_cache is None:
        with _safe_cache_lock:
            if _safe_global_cache is None:
                _safe_global_cache = SafeWorkbookCache(max_size=12, max_age_seconds=600)
    
    return _safe_global_cache


def get_safe_cached_workbook(file_path, data_only=True):
    """
    便捷函數：使用安全快取獲取工作簿
    
    Args:
        file_path: Excel檔案路徑
        data_only: 是否只讀取計算值
        
    Returns:
        openpyxl.Workbook: 工作簿物件
    """
    cache = get_safe_global_cache()
    return cache.get_workbook(file_path, data_only)


def clear_safe_cache():
    """清空安全快取"""
    global _safe_global_cache
    if _safe_global_cache is not None:
        _safe_global_cache.clear()


def print_safe_cache_stats():
    """打印安全快取統計"""
    cache = get_safe_global_cache()
    stats = cache.get_stats()
    
    print("\n=== Safe Workbook Cache Statistics ===")
    print(f"Cache Size: {stats['cache_size']}/{stats['max_size']}")
    print(f"Hit Rate: {stats['hit_rate_percent']}%")
    print(f"Hits: {stats['stats']['hits']}, Misses: {stats['stats']['misses']}")
    print(f"Evictions: {stats['stats']['evictions']}, Errors: {stats['stats']['errors']}")
    print(f"Memory Cleanups: {stats['stats']['memory_cleanups']}")
    if stats['cached_files']:
        print(f"Cached Files: {', '.join(stats['cached_files'])}")
    print("=====================================\n")


if __name__ == "__main__":
    # 簡單測試
    cache = SafeWorkbookCache(max_size=3)
    print("Safe cache system initialized successfully!")
    print_safe_cache_stats()