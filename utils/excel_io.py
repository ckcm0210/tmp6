# -*- coding: utf-8 -*-
"""
Excel I/O Utilities Module

This module contains functions for reading and writing Excel files,
including external file access and worksheet management.
"""

import os
import re
import openpyxl
import xlrd


def read_external_cell_value(current_workbook_path, external_file_full_path, external_sheet_name, cell_address):
    """
    Read cell value from an external Excel file.
    
    Args:
        current_workbook_path (str): Path to the current workbook
        external_file_full_path (str): Full path to the external file
        external_sheet_name (str): Name of the worksheet in external file
        cell_address (str): Cell address to read (e.g., 'A1')
        
    Returns:
        str: Formatted string containing the cell value or error message
    """
    full_external_path_normalized = os.path.normpath(external_file_full_path)
    if not os.path.exists(full_external_path_normalized):
        return f"External (File Not Found on Disk: {full_external_path_normalized})"
    
    file_extension = os.path.splitext(full_external_path_normalized)[1].lower()
    
    if file_extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        try:
            workbook = openpyxl.load_workbook(full_external_path_normalized, data_only=True, read_only=True)
            found_sheet = None
            for sname in workbook.sheetnames:
                if sname.lower() == external_sheet_name.lower():
                    found_sheet = sname
                    break
            if found_sheet:
                worksheet = workbook[found_sheet]
                cell_value = worksheet[cell_address].value
                workbook.close()
                return f"External (OpenPyxl): {cell_value if cell_value is not None else 'Empty'}"
            else:
                workbook.close()
                return "External (Sheet Not Found in file)"
        except Exception as e:
            return f"External (OpenPyxl Error: {str(e)[:100]})"
    
    if file_extension == '.xls':
        try:
            import xlrd
            workbook = xlrd.open_workbook(full_external_path_normalized, on_demand=True)
            found_sheet = None
            for sname in workbook.sheet_names():
                if sname.lower() == external_sheet_name.lower():
                    found_sheet = sname
                    break
            if found_sheet:
                worksheet = workbook.sheet_by_name(found_sheet)
                m = re.match(r'^([A-Z]+)([0-9]+)$', cell_address.replace('$', ''))
                if m:
                    col_letters, row_str = m.groups()
                    col_idx = 0
                    for i, c in enumerate(reversed(col_letters)):
                        col_idx += (ord(c.upper()) - ord('A') + 1) * (26 ** i)
                    col_idx -= 1
                    row_idx = int(row_str) - 1
                    if 0 <= row_idx < worksheet.nrows and 0 <= col_idx < worksheet.ncols:
                        cell_value = worksheet.cell_value(row_idx, col_idx)
                        return f"External (xlrd): {cell_value if cell_value != '' else 'Empty'}"
                    else:
                        return "External (Cell Address Out of Range)"
                else:
                    return "External (Invalid Cell Address Format)"
            else:
                return "External (Sheet Not Found in file)"
        except Exception as e:
            return f"External (xlrd Error: {str(e)[:100]})"
    
    return "External (Live reading for this file type is disabled)"


def find_matching_sheet(workbook, sheet_name):
    """
    Find a worksheet by name in a COM workbook object.
    
    Args:
        workbook: COM workbook object
        sheet_name (str): Name of the worksheet to find
        
    Returns:
        COM worksheet object or None if not found
    """
    try:
        for ws in workbook.Worksheets:
            if ws.Name == sheet_name:
                return ws
    except Exception as e:
        print(f"ERROR: Failed to get worksheet names: {e}")
    return None


def get_sheet_by_name(wb, sheet_name):
    """
    Get a worksheet by name from an openpyxl workbook.
    
    Args:
        wb: openpyxl workbook object
        sheet_name (str): Name of the worksheet
        
    Returns:
        openpyxl worksheet object
        
    Raises:
        ValueError: If worksheet is not found
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found in this workbook!")
    return wb[sheet_name]


def calculate_similarity(str1, str2):
    """
    Calculate similarity between two strings using edit distance.
    
    Args:
        str1 (str): First string
        str2 (str): Second string
        
    Returns:
        float: Similarity score between 0.0 and 1.0
    """
    len1, len2 = len(str1), len(str2)
    if len1 == 0 or len2 == 0:
        return 0.0
    
    dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]
    for i in range(len1 + 1):
        dp[i][0] = i
    for j in range(len2 + 1):
        dp[0][j] = j
    
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if str1[i-1] == str2[j-1]:
                dp[i][j] = dp[i-1][j-1]
            else:
                dp[i][j] = min(dp[i-1][j], dp[i][j-1], dp[i-1][j-1]) + 1
    
    edit_distance = dp[len1][len2]
    max_len = max(len1, len2)
    similarity = 1.0 - (edit_distance / max_len)
    return similarity