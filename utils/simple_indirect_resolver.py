# -*- coding: utf-8 -*-
"""
Simple INDIRECT Resolver - 借鑒unified_indirect_resolver的邏輯
只處理INDIRECT函數部分，將其還原為實際引用
"""

import re
import os
import openpyxl
from urllib.parse import unquote
import win32com.client as win32

class SimpleIndirectResolver:
    """簡單的INDIRECT解析器 - 只處理INDIRECT函數替換"""
    
    def __init__(self, workbook_path, sheet_name):
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.external_links_map = {}
        self.load_workbook()
    
    def load_workbook(self):
        """載入工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.workbook_path, data_only=False)
            self.worksheet = self.workbook[self.sheet_name]
            self.get_external_links()
        except Exception as e:
            print(f"Error loading workbook: {e}")
    
    def get_external_links(self):
        """獲取外部連結映射 - 借鑒你的邏輯"""
        try:
            self.external_links_map = {}
            
            if hasattr(self.workbook, '_external_links'):
                external_links = self.workbook._external_links
                if external_links:
                    for i, link in enumerate(external_links, 1):
                        if hasattr(link, 'file_link') and link.file_link:
                            file_path = link.file_link.Target
                            if file_path:
                                decoded_path = unquote(file_path)
                                if decoded_path.startswith('file:///'):
                                    decoded_path = decoded_path[8:]
                                elif decoded_path.startswith('file://'):
                                    decoded_path = decoded_path[7:]
                                
                                self.external_links_map[str(i)] = decoded_path
            
            # 如果沒有找到，推斷常見的外部連結
            if not self.external_links_map:
                self.infer_external_links()
                
        except Exception as e:
            self.infer_external_links()
    
    def infer_external_links(self):
        """推斷外部連結"""
        try:
            base_dir = os.path.dirname(self.workbook_path)
            common_files = [
                "Link1.xlsx", "Link2.xlsx", "Link3.xlsx",
                "File1.xlsx", "File2.xlsx", "File3.xlsx",
                "Data.xlsx", "GDP.xlsx", "Test.xlsx"
            ]
            
            index = 1
            for filename in common_files:
                full_path = os.path.join(base_dir, filename)
                if os.path.exists(full_path):
                    self.external_links_map[str(index)] = full_path
                    index += 1
                    
        except Exception as e:
            print(f"Error inferring external links: {e}")
    
    def find_indirect_functions(self, formula):
        """找到公式中的所有INDIRECT函數"""
        if not formula or not isinstance(formula, str):
            return []
        
        indirect_functions = []
        formula_upper = formula.upper()
        
        # 找到所有INDIRECT函數的位置
        start_pos = 0
        while True:
            indirect_pos = formula_upper.find('INDIRECT(', start_pos)
            if indirect_pos == -1:
                break
            
            # 提取完整的INDIRECT函數
            indirect_content, end_pos = self.extract_indirect_function(formula, indirect_pos)
            if indirect_content:
                indirect_functions.append({
                    'start_pos': indirect_pos,
                    'end_pos': end_pos,
                    'full_function': indirect_content,
                    'content': indirect_content[9:-1]  # 移除INDIRECT()
                })
            
            start_pos = indirect_pos + 1
        
        return indirect_functions
    
    def extract_indirect_function(self, formula, start_pos):
        """提取完整的INDIRECT函數"""
        try:
            # 從INDIRECT(後開始計算括號
            bracket_start = start_pos + len('INDIRECT(')
            bracket_count = 1
            current_pos = bracket_start
            in_quotes = False
            quote_char = None
            
            # 逐字符掃描，計算括號配對
            while current_pos < len(formula) and bracket_count > 0:
                char = formula[current_pos]
                
                # 處理引號
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                
                # 只在不在引號內時計算括號
                if not in_quotes:
                    if char == '(':
                        bracket_count += 1
                    elif char == ')':
                        bracket_count -= 1
                
                current_pos += 1
            
            if bracket_count == 0:
                full_function = formula[start_pos:current_pos]
                return full_function, current_pos
            
            return None, start_pos
        except Exception as e:
            print(f"Error extracting INDIRECT function: {e}")
            return None, start_pos
    
    def resolve_indirect_content(self, indirect_content, context_cell=None):
        """解析INDIRECT內容 - 使用Excel COM安全計算"""
        try:
            print(f"Resolving INDIRECT content: {indirect_content}")
            
            # 嘗試使用Excel COM進行安全計算
            excel_result = self.safe_excel_calculation(indirect_content, context_cell)
            if excel_result is not None:
                return excel_result
            
            # 如果Excel COM失敗，回退到純解析模式
            print("Excel COM failed, falling back to pure mode")
            return self.pure_mode_calculation(indirect_content, context_cell)
                
        except Exception as e:
            print(f"Error resolving INDIRECT content: {e}")
            return None
    
    def safe_excel_calculation(self, indirect_content, context_cell=None):
        """安全的Excel計算方法 - 借鑒你的邏輯"""
        try:
            # 檢查是否有Excel COM
            try:
                import win32com.client as win32
            except ImportError:
                print("Excel COM not available")
                return None
            
            print("Starting safe Excel calculation...")
            
            # 連接Excel
            try:
                xl = win32.GetActiveObject("Excel.Application")
            except:
                try:
                    xl = win32.Dispatch("Excel.Application")
                    xl.Visible = False
                except:
                    print("Cannot connect to Excel")
                    return None
            
            # 打開工作簿
            excel_workbook = xl.Workbooks.Open(self.workbook_path)
            excel_ws = excel_workbook.Worksheets(self.sheet_name)
            
            # 找一個空白儲存格進行計算
            test_cell = excel_ws.Range("ZZ1")  # 使用ZZ1作為測試儲存格
            
            # 保存原始狀態
            original_formula = test_cell.Formula
            original_calculation = xl.Calculation
            original_events = xl.EnableEvents
            original_screen_updating = xl.ScreenUpdating
            original_interactive = xl.Interactive
            
            try:
                # 設置保護模式 - 借鑒你的邏輯
                xl.Calculation = -4135  # xlCalculationManual
                xl.EnableEvents = False
                xl.ScreenUpdating = False
                xl.Interactive = False
                
                print("Protection mode activated")
                
                # 在Excel中計算INDIRECT
                test_formula = f"=INDIRECT({indirect_content})"
                test_cell.Formula = test_formula
                test_cell.Calculate()
                
                # 獲取計算結果
                result_value = test_cell.Value
                
                # 如果結果是地址，轉換為字串
                if result_value is not None:
                    # 嘗試獲取地址字串
                    try:
                        address_result = test_cell.Formula
                        if address_result and address_result.startswith('='):
                            # 獲取實際的引用地址
                            calculated_address = str(result_value)
                            print(f"Excel calculation result: {calculated_address}")
                            return calculated_address
                    except:
                        pass
                
                print(f"Excel calculation result: {result_value}")
                return str(result_value) if result_value is not None else None
                
            finally:
                # 恢復所有狀態 - 借鑒你的邏輯
                test_cell.Formula = original_formula
                xl.Calculation = original_calculation
                xl.EnableEvents = original_events
                xl.ScreenUpdating = original_screen_updating
                xl.Interactive = original_interactive
                
                # 關閉工作簿（不保存）
                excel_workbook.Close(SaveChanges=False)
                
                print("All states restored")
                
        except Exception as e:
            print(f"Error in safe Excel calculation: {e}")
            return None
    
    def pure_mode_calculation(self, indirect_content, context_cell=None):
        """純模式計算 - 當Excel COM不可用時的回退方案"""
        try:
            print("Using pure mode calculation...")
            
            # 修復外部引用
            fixed_content = self.fix_external_references(indirect_content)
            
            # 解析字串連接
            if '&' in fixed_content:
                result = self.resolve_concatenation(fixed_content, context_cell)
                return result
            else:
                # 簡單引用，移除引號
                if fixed_content.startswith('"') and fixed_content.endswith('"'):
                    fixed_content = fixed_content[1:-1]
                return fixed_content
                
        except Exception as e:
            print(f"Error in pure mode calculation: {e}")
            return None
    
    def fix_external_references(self, content):
        """修復外部引用 - 借鑒你的邏輯"""
        try:
            def replace_ref(match):
                ref_num = match.group(1)
                if ref_num in self.external_links_map:
                    full_path = self.external_links_map[ref_num]
                    decoded_path = unquote(full_path) if isinstance(full_path, str) else full_path
                    if decoded_path.startswith('file:///'):
                        decoded_path = decoded_path[8:]
                    
                    filename = os.path.basename(decoded_path)
                    directory = os.path.dirname(decoded_path)
                    return f"'{directory}\\[{filename}]'"
                return f"[Unknown_{ref_num}]"
            
            pattern = r'\[(\d+)\]'
            return re.sub(pattern, replace_ref, content)
        except:
            return content
    
    def resolve_concatenation(self, content, context_cell=None):
        """解析字串連接 - 借鑒你的邏輯"""
        try:
            # 按 & 分割（智能處理引號內的&）
            parts = self.smart_split_by_ampersand(content)
            
            result_parts = []
            for part in parts:
                part = part.strip()
                
                # 字串常數
                if (part.startswith('"') and part.endswith('"')) or \
                   (part.startswith("'") and part.endswith("'")):
                    result_parts.append(part[1:-1])
                
                # 儲存格引用
                elif re.match(r'^\$?[A-Z]+\$?\d+$', part):
                    cell_value = self.get_cell_value(part)
                    result_parts.append(str(cell_value) if cell_value is not None else "")
                
                # ROW()函數
                elif 'ROW()' in part.upper() and context_cell:
                    row_num = int(re.search(r'\d+', context_cell).group())
                    if '+' in part:
                        match = re.search(r'ROW\(\)\s*\+\s*(\d+)', part, re.IGNORECASE)
                        if match:
                            add_num = int(match.group(1))
                            result_parts.append(str(row_num + add_num))
                        else:
                            result_parts.append(str(row_num))
                    else:
                        result_parts.append(str(row_num))
                
                # COLUMN()函數
                elif 'COLUMN()' in part.upper() and context_cell:
                    col_letters = re.search(r'[A-Z]+', context_cell).group()
                    col_num = 0
                    for char in col_letters:
                        col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                    result_parts.append(str(col_num))
                
                else:
                    # 其他，保持原樣
                    result_parts.append(part)
            
            final_result = ''.join(result_parts)
            print(f"Concatenation result: {final_result}")
            return final_result
            
        except Exception as e:
            print(f"Error in concatenation: {e}")
            return content
    
    def smart_split_by_ampersand(self, content):
        """按 & 分割，但不會分割引號內的 &"""
        try:
            parts = []
            current_part = ""
            in_quotes = False
            quote_char = None
            
            i = 0
            while i < len(content):
                char = content[i]
                
                # 處理引號
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                    current_part += char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                    current_part += char
                elif char == '&' and not in_quotes:
                    # 分割點
                    if current_part.strip():
                        parts.append(current_part.strip())
                    current_part = ""
                else:
                    current_part += char
                
                i += 1
            
            # 加最後一部分
            if current_part.strip():
                parts.append(current_part.strip())
            
            return parts
        except Exception as e:
            print(f"Error in smart split: {e}")
            return [content]
    
    def get_cell_value(self, cell_ref):
        """獲取儲存格值"""
        try:
            cell = self.worksheet[cell_ref]
            return cell.value
        except:
            return None
    
    def resolve_formula_indirect(self, formula, context_cell=None):
        """
        解析公式中的INDIRECT函數，返回resolved formula
        
        例如：
        輸入：=INDIRECT(A1&"!"&B10) + A2
        輸出：=工作表1!B10 + A2
        """
        if not formula or 'INDIRECT' not in formula.upper():
            return formula
        
        try:
            print(f"Resolving formula: {formula}")
            
            # 找到所有INDIRECT函數
            indirect_functions = self.find_indirect_functions(formula)
            
            if not indirect_functions:
                return formula
            
            # 從後往前替換，避免位置偏移
            resolved_formula = formula
            for indirect_func in reversed(indirect_functions):
                # 解析INDIRECT內容
                resolved_ref = self.resolve_indirect_content(
                    indirect_func['content'], 
                    context_cell
                )
                
                if resolved_ref:
                    # 替換INDIRECT函數為解析後的引用
                    resolved_formula = (
                        resolved_formula[:indirect_func['start_pos']] + 
                        resolved_ref + 
                        resolved_formula[indirect_func['end_pos']:]
                    )
                    print(f"Replaced {indirect_func['full_function']} with {resolved_ref}")
            
            print(f"Final resolved formula: {resolved_formula}")
            return resolved_formula
            
        except Exception as e:
            print(f"Error resolving formula INDIRECT: {e}")
            return formula


def resolve_indirect_in_formula(formula, workbook_path, sheet_name, context_cell=None):
    """
    便捷函數：解析公式中的INDIRECT函數
    
    Args:
        formula: 包含INDIRECT的公式
        workbook_path: Excel文件路徑
        sheet_name: 工作表名稱
        context_cell: 公式所在的儲存格地址
        
    Returns:
        dict: {
            'has_indirect': bool,
            'original_formula': str,
            'resolved_formula': str
        }
    """
    try:
        if not formula or 'INDIRECT' not in formula.upper():
            return {
                'has_indirect': False,
                'original_formula': formula,
                'resolved_formula': formula
            }
        
        resolver = SimpleIndirectResolver(workbook_path, sheet_name)
        resolved_formula = resolver.resolve_formula_indirect(formula, context_cell)
        
        return {
            'has_indirect': True,
            'original_formula': formula,
            'resolved_formula': resolved_formula
        }
        
    except Exception as e:
        print(f"Error in resolve_indirect_in_formula: {e}")
        return {
            'has_indirect': False,
            'original_formula': formula,
            'resolved_formula': formula
        }


# 測試函數
if __name__ == "__main__":
    # 測試用例
    test_formula = '=INDIRECT(A1&"!"&B10) + A2'
    print(f"Test formula: {test_formula}")
    
    # 這裡需要實際的Excel文件來測試
    print("Simple INDIRECT Resolver ready for integration!")