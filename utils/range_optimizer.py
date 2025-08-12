import re
import collections
from openpyxl.utils import column_index_from_string, get_column_letter

def parse_excel_address(addr):
    """
    Parse Excel address and return type and normalized format.
    Supports: single cells, ranges, row ranges, column ranges.
    """
    addr = addr.replace('$', '').strip().upper()

    if not addr:
        raise ValueError("Address input cannot be empty.")

    if re.fullmatch(r"^[0-9]+(:[0-9]+)?$", addr):
        parts = list(map(int, addr.split(':')))
        start, end = (parts[0], parts[0]) if len(parts) == 1 else (parts[0], parts[1])
        if start > end:
            start, end = end, start
        return ('row_range', f"{start}:{end}")

    if re.fullmatch(r"^[A-Z]+(:[A-Z]+)?$", addr):
        parts = addr.split(':')
        start_col, end_col = (parts[0], parts[0]) if len(parts) == 1 else (parts[0], parts[1])
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
        start_col_sorted = get_column_letter(start_idx)
        end_col_sorted = get_column_letter(end_idx)
        return ('col_range', f"{start_col_sorted}:{end_col_sorted}")

    if re.fullmatch(r"^[A-Z]+[0-9]+$", addr):
        return ('cell', addr)
        
    m = re.fullmatch(r"^([A-Z]+[0-9]+):([A-Z]+[0-9]+)$", addr)
    if m:
        c1, c2 = m.groups()
        c1_col_str, c1_row_str = re.match(r"([A-Z]+)([0-9]+)", c1).groups()
        c2_col_str, c2_row_str = re.match(r"([A-Z]+)([0-9]+)", c2).groups()
        
        c1_col = column_index_from_string(c1_col_str)
        c2_col = column_index_from_string(c2_col_str)
        c1_row = int(c1_row_str)
        c2_row = int(c2_row_str)

        start_col_idx = min(c1_col, c2_col)
        end_col_idx = max(c1_col, c2_col)
        start_row = min(c1_row, c2_row)
        end_row = max(c1_row, c2_row)
        
        start_cell = f"{get_column_letter(start_col_idx)}{start_row}"
        end_cell = f"{get_column_letter(end_col_idx)}{end_row}"
        
        return ('range', f"{start_cell}:{end_cell}")

    raise ValueError(f"Invalid address format: '{addr}'")

def parse_cell_address(addr):
    match = re.match(r'([A-Z]+)(\d+)', addr.upper())
    if match:
        col_str, row_str = match.groups()
        col_num = 0
        for char in col_str:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        return (col_num, int(row_str))
    return None

def format_range(start_addr, end_addr):
    if start_addr == end_addr:
        return start_addr
    return f"{start_addr}:{end_addr}"

def optimize_ranges(parsed_addresses):
    if not parsed_addresses:
        return []
    # Simple consecutive grouping for small sets
    if len(parsed_addresses) <= 3:
        ranges = []
        current_start = parsed_addresses[0][1]
        current_end = parsed_addresses[0][1]
        for i in range(1, len(parsed_addresses)):
            prev_col, prev_row = parsed_addresses[i-1][0]
            curr_col, curr_row = parsed_addresses[i][0]
            if ((prev_col == curr_col and curr_row == prev_row + 1) or 
                (prev_row == curr_row and curr_col == prev_col + 1)):
                current_end = parsed_addresses[i][1]
            else:
                ranges.append(format_range(current_start, current_end))
                current_start = parsed_addresses[i][1]
                current_end = parsed_addresses[i][1]
        ranges.append(format_range(current_start, current_end))
        return ranges

    # For larger sets, try rectangle detection
    def detect_rectangles(addresses):
        col_ranges = collections.defaultdict(list)
        for (col, row), addr in addresses:
            col_ranges[col].append(row)
        for col in col_ranges:
            col_ranges[col].sort()
        
        rectangles = []
        used_addresses = set()
        cols = sorted(col_ranges.keys())
        for start_col in cols:
            for end_col in cols[cols.index(start_col):]:
                common_rows = set(col_ranges[start_col])
                for col in range(start_col + 1, end_col + 1):
                    if col in col_ranges:
                        common_rows &= set(col_ranges[col])
                
                if len(common_rows) >= 2:
                    sorted_rows = sorted(common_rows)
                    for i in range(len(sorted_rows)):
                        for j in range(i + 1, len(sorted_rows) + 1):
                            row_range = sorted_rows[i:j]
                            if len(row_range) >= 2 and row_range == list(range(row_range[0], row_range[-1] + 1)):
                                rect_addresses = set()
                                for col in range(start_col, end_col + 1):
                                    for row in row_range:
                                        for (c, r), addr in addresses:
                                            if c == col and r == row:
                                                rect_addresses.add(((c, r), addr))
                                
                                if len(rect_addresses) > 1 and not rect_addresses & used_addresses:
                                    start_addr, end_addr = None, None
                                    for (c, r), addr in rect_addresses:
                                        if start_addr is None:
                                            start_addr, end_addr = addr, addr
                                        else:
                                            end_addr = addr
                                    
                                    if start_col == end_col:
                                        rect_range_str = f"{chr(ord('A') + start_col - 1)}{row_range[0]}:{chr(ord('A') + start_col - 1)}{row_range[-1]}"
                                    else:
                                        rect_range_str = f"{chr(ord('A') + start_col - 1)}{row_range[0]}:{chr(ord('A') + end_col - 1)}{row_range[-1]}"
                                    
                                    rectangles.append((len(rect_addresses), rect_range_str, rect_addresses))
                                    used_addresses.update(rect_addresses)
        
        rectangles.sort(key=lambda x: x[0], reverse=True)
        if rectangles:
            best_rect = rectangles[0]
            remaining = [addr for addr_info, addr in addresses if addr_info not in {addr_info for addr_info, addr in best_rect[2]}]
            return [best_rect[1]], remaining
        return [], [addr for addr_info, addr in addresses]

    rect_ranges, remaining_addrs = detect_rectangles(parsed_addresses)
    if remaining_addrs:
        remaining_parsed = []
        for addr in remaining_addrs:
            parsed = parse_cell_address(addr)
            if parsed:
                remaining_parsed.append((parsed, addr))
        remaining_parsed.sort(key=lambda x: x[0])
        
        if remaining_parsed:
            current_start = remaining_parsed[0][1]
            current_end = remaining_parsed[0][1]
            for i in range(1, len(remaining_parsed)):
                prev_col, prev_row = remaining_parsed[i-1][0]
                curr_col, curr_row = remaining_parsed[i][0]
                if ((prev_col == curr_col and curr_row == prev_row + 1) or 
                    (prev_row == curr_row and curr_col == prev_col + 1)):
                    current_end = remaining_parsed[i][1]
                else:
                    rect_ranges.append(format_range(current_start, current_end))
                    current_start = remaining_parsed[i][1]
                    current_end = remaining_parsed[i][1]
            rect_ranges.append(format_range(current_start, current_end))
    return rect_ranges

def smart_range_display(addresses):
    if not addresses:
        return ""
    parsed = [p for p in (parse_cell_address(addr) for addr in addresses) if p]
    if not parsed:
        return f"{len(addresses)} cells"
    
    parsed_with_addr = sorted([(p, addr) for p, addr in zip(parsed, addresses)])
    
    # This is a simplified version for display, can be enhanced
    ranges = optimize_ranges(parsed_with_addr)
    
    if len(ranges) <= 8:
        return f"{len(addresses)} cells: {', '.join(ranges)}"
    else:
        sample_ranges = ranges[:5]
        return f"{len(addresses)} cells: {', '.join(sample_ranges)}, ... and {len(ranges)-5} more ranges"