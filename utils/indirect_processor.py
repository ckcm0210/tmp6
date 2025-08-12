# -*- coding: utf-8 -*-
"""
INDIRECT Function Processor - 處理INDIRECT函數解析
整合unified_indirect_resolver的核心功能到dependency exploder中
"""

import re
import os
from urllib.parse import unquote
import openpyxl
import sys
import traceback

class IndirectProcessor:
    """INDIRECT函數處理器"""
    
    def __init__(self, workbook_path, sheet_name):
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.worksheet = None
        self.external_links_map = {}
        self.load_workbook()
    
    def load_workbook(self):
        """載入工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.workbook_path, data_only=False)
            self.worksheet = self.workbook[self.sheet_name]
            self.get_external_links_from_openpyxl()
        except Exception as e:
            print(f"Error loading workbook: {e}")
    
    def identify_indirect_functions(self, formula):
        """識別公式中的INDIRECT函數"""
        if not formula or not isinstance(formula, str):
            return []
        
        indirect_functions = []
        formula_upper = formula.upper()
        
        # 找到所有INDIRECT函數的位置
        start_pos = 0
        while True:
            indirect_pos = formula_upper.find('INDIRECT(', start_pos)
            if indirect_pos == -1:
                break
            
            # 提取完整的INDIRECT函數
            indirect_content = self.extract_indirect_content(formula, indirect_pos)
            if indirect_content:
                indirect_functions.append({
                    'position': indirect_pos,
                    'function': f"INDIRECT({indirect_content})",
                    'content': indirect_content,
                    'original': f"INDIRECT({indirect_content})"
                })
            
            start_pos = indirect_pos + 1
        
        return indirect_functions
    
    def extract_indirect_content(self, formula, start_pos):
        """提取INDIRECT函數的內容"""
        try:
            # 從INDIRECT(後開始計算括號
            bracket_start = start_pos + len('INDIRECT(')
            bracket_count = 1
            current_pos = bracket_start
            in_quotes = False
            quote_char = None
            
            # 逐字符掃描，計算括號配對，處理引號
            while current_pos < len(formula) and bracket_count > 0:
                char = formula[current_pos]
                
                # 處理引號（單引號和雙引號）
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                
                # 只在不在引號內時計算括號
                if not in_quotes:
                    if char == '(':
                        bracket_count += 1
                    elif char == ')':
                        bracket_count -= 1
                
                current_pos += 1
            
            if bracket_count == 0:
                return formula[bracket_start:current_pos-1]
            
            return None
        except Exception as e:
            print(f"Error extracting INDIRECT content: {e}")
            return None
    
    def resolve_indirect_function(self, indirect_content, context_cell=None):
        """解析INDIRECT函數，返回最終的引用地址"""
        try:
            print(f"Resolving INDIRECT: {indirect_content}")
            
            # 修復外部引用
            fixed_content = self.fix_external_references(indirect_content)
            print(f"After external reference fix: {fixed_content}")
            
            # 解析字串連接
            if '&' in fixed_content:
                print("Contains string concatenation - resolving components")
                components = self.resolve_concatenation_components(fixed_content, context_cell)
                result = self.build_final_reference(components)
                return result
            else:
                print("Simple reference - no concatenation")
                # 移除引號
                if fixed_content.startswith('"') and fixed_content.endswith('"'):
                    fixed_content = fixed_content[1:-1]
                return fixed_content
                
        except Exception as e:
            print(f"Error resolving INDIRECT: {e}")
            return None
    
    def get_external_links_from_openpyxl(self):
        """從openpyxl獲取外部連結映射"""
        try:
            self.external_links_map = {}
            
            if hasattr(self.workbook, '_external_links'):
                external_links = self.workbook._external_links
                if external_links:
                    for i, link in enumerate(external_links, 1):
                        if hasattr(link, 'file_link') and link.file_link:
                            file_path = link.file_link.Target
                            if file_path:
                                decoded_path = unquote(file_path)
                                if decoded_path.startswith('file:///'):
                                    decoded_path = decoded_path[8:]
                                elif decoded_path.startswith('file://'):
                                    decoded_path = decoded_path[7:]
                                
                                self.external_links_map[str(i)] = decoded_path
                                print(f"External link [{i}] = {os.path.basename(decoded_path)}")
            
            if not self.external_links_map:
                self.infer_external_links_from_formulas()
                
        except Exception as e:
            self.infer_external_links_from_formulas()
    
    def infer_external_links_from_formulas(self):
        """從公式中推斷外部連結"""
        try:
            base_dir = os.path.dirname(self.workbook_path)
            common_files = [
                "Link1.xlsx", "Link2.xlsx", "Link3.xlsx",
                "File1.xlsx", "File2.xlsx", "File3.xlsx",
                "Data.xlsx", "GDP.xlsx", "Test.xlsx"
            ]
            
            index = 1
            for filename in common_files:
                full_path = os.path.join(base_dir, filename)
                if os.path.exists(full_path):
                    self.external_links_map[str(index)] = full_path
                    print(f"Inferred link [{index}] = {filename}")
                    index += 1
                    
        except Exception as e:
            print(f"Error inferring external links: {e}")
    
    def fix_external_references(self, content):
        """修復外部引用"""
        try:
            def replace_ref(match):
                ref_num = match.group(1)
                if ref_num in self.external_links_map:
                    full_path = self.external_links_map[ref_num]
                    decoded_path = unquote(full_path) if isinstance(full_path, str) else full_path
                    if decoded_path.startswith('file:///'):
                        decoded_path = decoded_path[8:]
                    
                    filename = os.path.basename(decoded_path)
                    directory = os.path.dirname(decoded_path)
                    return f"'{directory}\\[{filename}]'"
                return f"[Unknown_{ref_num}]"
            
            pattern = r'\[(\d+)\]'
            return re.sub(pattern, replace_ref, content)
        except:
            return content
    
    def resolve_concatenation_components(self, content, context_cell=None):
        """解析字串連接組件"""
        try:
            print("Starting component analysis...")
            
            # 按 & 分割（智能處理引號內的&）
            parts = self.smart_split_by_ampersand(content)
            print(f"Split into {len(parts)} parts: {parts}")
            
            components = []
            
            for part in parts:
                part = part.strip()
                print(f"Analyzing component: {part}")
                
                # 識別組件類型
                comp_type, comp_data = self.identify_component_type(part)
                print(f"  Type: {comp_type}")
                
                if comp_type == 'string':
                    components.append(('string', comp_data))
                    print(f"  String constant: '{comp_data}'")
                
                elif comp_type == 'cell':
                    cell_value = self.get_cell_value_with_formula_calc(comp_data, context_cell)
                    components.append(('cell', comp_data, cell_value))
                    print(f"  Cell {comp_data} = {cell_value}")
                
                elif comp_type == 'function':
                    func_result = self.resolve_function_smart(part, context_cell)
                    components.append(('function', part, func_result))
                    print(f"  Function result = {func_result}")
                
                else:
                    components.append(('expression', part, None))
                    print(f"  Complex expression: {part}")
            
            return components
        except Exception as e:
            print(f"Error in component analysis: {e}")
            return []
    
    def build_final_reference(self, components):
        """構建最終引用"""
        try:
            print("Building final reference from components:")
            
            result_parts = []
            for component in components:
                comp_type = component[0]
                comp_value = component[2] if len(component) > 2 else None
                
                if comp_type == 'string':
                    result_parts.append(component[1])
                elif comp_type in ['cell', 'function'] and comp_value is not None:
                    result_parts.append(str(comp_value))
                else:
                    result_parts.append(f"({component[1]})")
            
            final_reference = ''.join(result_parts)
            print(f"Final reference: {final_reference}")
            
            return final_reference
        except Exception as e:
            print(f"Error building final reference: {e}")
            return None
    
    def smart_split_by_ampersand(self, content):
        """按 & 分割，但不會分割引號內的 &"""
        try:
            parts = []
            current_part = ""
            in_quotes = False
            quote_char = None
            
            i = 0
            while i < len(content):
                char = content[i]
                
                # 處理引號
                if char in ['"', "'"] and not in_quotes:
                    in_quotes = True
                    quote_char = char
                    current_part += char
                elif char == quote_char and in_quotes:
                    in_quotes = False
                    quote_char = None
                    current_part += char
                elif char == '&' and not in_quotes:
                    # 分割點
                    if current_part.strip():
                        parts.append(current_part.strip())
                    current_part = ""
                else:
                    current_part += char
                
                i += 1
            
            # 加最後一部分
            if current_part.strip():
                parts.append(current_part.strip())
            
            return parts
        except Exception as e:
            print(f"Error in smart split: {e}")
            return [content]
    
    def identify_component_type(self, component):
        """識別組件類型"""
        try:
            comp = component.strip()
            
            # 1. 字串常數 (用引號包住)
            if (comp.startswith('"') and comp.endswith('"')) or \
               (comp.startswith("'") and comp.endswith("'")):
                return ('string', comp[1:-1])
            
            # 2. 簡單儲存格引用 (例如: B8, $A$1)
            if re.match(r'^\$?[A-Z]+\$?\d+$', comp):
                return ('cell', comp)
            
            # 3. 函數 (有英文字母 + 開括弧)
            if re.match(r'^[A-Z]+\s*\(', comp):
                return ('function', comp)
            
            # 4. 其他表達式
            return ('expression', comp)
        except:
            return ('expression', component)
    
    def get_cell_value_with_formula_calc(self, cell_ref, context_cell=None):
        """獲取儲存格值，包含公式計算"""
        try:
            cell = self.worksheet[cell_ref]
            raw_value = cell.value
            
            # 如果不是公式，直接返回
            if not (isinstance(raw_value, str) and raw_value.startswith('=')):
                return raw_value
            
            # 是公式，嘗試計算
            print(f"    Found formula in {cell_ref}: {raw_value}")
            
            # 移除開頭的=號
            formula = raw_value[1:]
            
            # 處理ROW()函數
            if 'ROW()' in formula.upper():
                result = self.resolve_position_aware_function(formula, 'ROW', context_cell or cell_ref)
                print(f"    ROW formula result: {result}")
                return result
            
            # 處理COLUMN()函數
            elif 'COLUMN()' in formula.upper():
                result = self.resolve_position_aware_function(formula, 'COLUMN', context_cell or cell_ref)
                print(f"    COLUMN formula result: {result}")
                return result
            
            # 處理字串連接
            elif '&' in formula:
                result = self.calculate_string_concat_formula(formula, context_cell or cell_ref)
                print(f"    String concat result: {result}")
                return result
            
            else:
                # 其他公式暫時返回原始值
                print(f"    Unknown formula type, returning original: {raw_value}")
                return raw_value
            
        except Exception as e:
            print(f"    Error calculating formula: {e}")
            return raw_value if 'raw_value' in locals() else None
    
    def resolve_position_aware_function(self, func_expr, func_type, context_cell):
        """解析位置相關函數（ROW, COLUMN等）"""
        try:
            if context_cell and func_type.upper() == 'ROW':
                row_num = int(re.search(r'\d+', context_cell).group())
                print(f"        ROW() context: {context_cell}, row: {row_num}")
                
                # 處理 ROW()+數字 的情況
                if '+' in func_expr:
                    match = re.search(r'ROW\(\)\s*\+\s*(\d+)', func_expr)
                    if match:
                        add_num = int(match.group(1))
                        result = row_num + add_num
                        print(f"        ROW()+{add_num} = {result}")
                        return result
                
                return row_num
            
            elif context_cell and func_type.upper() == 'COLUMN':
                col_letters = re.search(r'[A-Z]+', context_cell).group()
                col_num = 0
                for char in col_letters:
                    col_num = col_num * 26 + (ord(char) - ord('A') + 1)
                print(f"        COLUMN() context: {context_cell}, column: {col_num}")
                return col_num
            
            return f"{func_type}()"
        except:
            return f"{func_type}()"
    
    def calculate_string_concat_formula(self, formula, current_cell):
        """計算字串連接公式"""
        try:
            print(f"      Calculating string concat: {formula}")
            
            # 按 & 分割
            parts = self.smart_split_by_ampersand(formula)
            result_parts = []
            
            for part in parts:
                part = part.strip()
                print(f"      Processing part: {part}")
                
                if part.startswith('"') and part.endswith('"'):
                    # 字串常數
                    value = part[1:-1]
                    result_parts.append(value)
                    print(f"        String: '{value}'")
                
                elif re.match(r'^\$?[A-Z]+\$?\d+$', part):
                    # 儲存格引用
                    cell_value = self.worksheet[part].value
                    result_parts.append(str(cell_value) if cell_value is not None else "")
                    print(f"        Cell {part}: {cell_value}")
                
                elif 'ROW()' in part.upper():
                    # ROW函數
                    row_result = self.resolve_position_aware_function(part, 'ROW', current_cell)
                    result_parts.append(str(row_result))
                    print(f"        ROW function: {row_result}")
                
                elif 'COLUMN()' in part.upper():
                    # COLUMN函數
                    col_result = self.resolve_position_aware_function(part, 'COLUMN', current_cell)
                    result_parts.append(str(col_result))
                    print(f"        COLUMN function: {col_result}")
                
                else:
                    # 其他，保持原樣
                    result_parts.append(part)
                    print(f"        Other: {part}")
            
            final_result = ''.join(result_parts)
            print(f"      Final concat result: {final_result}")
            return final_result
            
        except Exception as e:
            print(f"      Error in string concat: {e}")
            return formula
    
    def resolve_function_smart(self, function_text, context_cell=None):
        """智能解析函數"""
        try:
            func_upper = function_text.upper()
            
            # ROW函數
            if func_upper.startswith('ROW'):
                return self.resolve_position_aware_function(function_text, 'ROW', context_cell)
            
            # COLUMN函數
            elif func_upper.startswith('COLUMN'):
                return self.resolve_position_aware_function(function_text, 'COLUMN', context_cell)
            
            # 其他函數暫時返回原始文字
            else:
                print(f"    Unknown function type: {function_text}")
                return function_text
                
        except Exception as e:
            print(f"    Error resolving function: {e}")
            return function_text
    
    def create_resolved_formula(self, original_formula, indirect_functions, resolved_references):
        """創建解析後的公式，將INDIRECT替換為實際引用"""
        try:
            resolved_formula = original_formula
            
            # 按位置從後往前替換，避免位置偏移
            replacements = []
            for i, indirect_func in enumerate(indirect_functions):
                if i < len(resolved_references) and resolved_references[i]:
                    replacements.append({
                        'original': indirect_func['original'],
                        'resolved': resolved_references[i]
                    })
            
            # 執行替換
            for replacement in replacements:
                resolved_formula = resolved_formula.replace(
                    replacement['original'], 
                    replacement['resolved']
                )
            
            return resolved_formula
        except Exception as e:
            print(f"Error creating resolved formula: {e}")
            return original_formula


def process_indirect_in_formula(formula, workbook_path, sheet_name, context_cell=None):
    """
    便捷函數：處理公式中的INDIRECT函數 - 使用你的unified_indirect_resolver
    
    Args:
        formula: 包含INDIRECT的公式
        workbook_path: Excel文件路徑
        sheet_name: 工作表名稱
        context_cell: 公式所在的儲存格地址（用於ROW/COLUMN函數）
        
    Returns:
        dict: {
            'has_indirect': bool,
            'indirect_functions': list,
            'resolved_references': list,
            'resolved_formula': str
        }
    """
    try:
        # 檢查是否包含INDIRECT
        if 'INDIRECT' not in formula.upper():
            return {
                'has_indirect': False,
                'indirect_functions': [],
                'resolved_references': [],
                'resolved_formula': formula
            }
        
        # === 使用你的unified_indirect_resolver ===
        import sys
        import os
        
        # 添加indirect_tool路徑
        indirect_tool_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'indirect_tool')
        if indirect_tool_path not in sys.path:
            sys.path.append(indirect_tool_path)
        
        from unified_indirect_resolver import UnifiedIndirectResolver
        
        print(f"Processing INDIRECT in formula: {formula}")
        print(f"Workbook: {workbook_path}")
        print(f"Sheet: {sheet_name}")
        print(f"Context cell: {context_cell}")
        
        # 創建resolver實例 - 無參數構造
        resolver = UnifiedIndirectResolver()
        
        # 設置工作簿和工作表
        resolver.workbook_path = workbook_path
        resolver.sheet_name = sheet_name
        
        # 設置當前儲存格（如果有context_cell）
        if context_cell:
            resolver.current_cell = context_cell
        
        # 設置要解析的公式
        resolver.formula = formula
        
        # 執行INDIRECT解析
        result = resolver.resolve_indirect_unified()
        
        print(f"Resolver result: {result}")
        
        if result and 'resolved_formula' in result:
            resolved_formula = result['resolved_formula']
            
            # 識別原始INDIRECT函數（簡單識別）
            indirect_functions = []
            if 'INDIRECT(' in formula.upper():
                # 簡單提取INDIRECT函數
                import re
                indirect_matches = re.findall(r'INDIRECT\([^)]+\)', formula, re.IGNORECASE)
                for match in indirect_matches:
                    indirect_functions.append({
                        'original': match,
                        'content': match[9:-1]  # 移除INDIRECT()
                    })
            
            return {
                'has_indirect': True,
                'indirect_functions': indirect_functions,
                'resolved_references': result.get('resolved_references', []),
                'resolved_formula': resolved_formula
            }
        else:
            print("Resolver returned no result")
            return {
                'has_indirect': True,
                'indirect_functions': [],
                'resolved_references': [],
                'resolved_formula': formula  # 保持原始公式
            }
        
    except Exception as e:
        print(f"Error processing INDIRECT in formula: {e}")
        import traceback
        traceback.print_exc()
        return {
            'has_indirect': False,
            'indirect_functions': [],
            'resolved_references': [],
            'resolved_formula': formula
        }


# 測試函數
if __name__ == "__main__":
    # 測試用例
    test_formula = '=INDIRECT("A"&ROW())'
    test_workbook = r"C:\Users\user\Desktop\pytest\test.xlsx"
    test_sheet = "Sheet1"
    test_cell = "B5"
    
    try:
        result = process_indirect_in_formula(test_formula, test_workbook, test_sheet, test_cell)
        print("INDIRECT Processing Result:")
        print(f"Has INDIRECT: {result['has_indirect']}")
        print(f"Original: {test_formula}")
        print(f"Resolved: {result['resolved_formula']}")
        print(f"References: {result['resolved_references']}")
    except Exception as e:
        print(f"Test failed: {e}")