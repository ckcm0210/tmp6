# -*- coding: utf-8 -*-
"""
Enhanced openpyxl with external reference resolution
Based on the provided ResolvedWorkbookView implementation
"""

import openpyxl
import os
import re
from .safe_cache import get_safe_cached_workbook
import traceback

# 輔助函數：從工作簿中獲取外部連結映射
def _get_external_link_map(workbook):
    external_link_map = {}
    if hasattr(workbook, '_external_links') and workbook._external_links:
        for i, link in enumerate(workbook._external_links):
            if hasattr(link, 'file_link') and hasattr(link.file_link, 'target'):
                target_path = link.file_link.target
                excel_formatted_path_part = ""
                if target_path.startswith('file:///'):
                    actual_path = target_path[len('file:///'):]
                    actual_path = actual_path.replace('\\', '\\\\')
                    actual_path = actual_path.replace('/', '\\\\')

                    dirname = os.path.dirname(actual_path)
                    basename = os.path.basename(actual_path)
                    # 修復格式：應該是 'path\[file.xlsx]sheet' 而不是 'path\[file.xlsx]'sheet
                    excel_formatted_path_part = f"'{dirname}\\\\[{basename}]"
                else:
                    excel_formatted_path_part = f"'[{target_path}]"
                external_link_map[str(i + 1)] = excel_formatted_path_part
    return external_link_map

# 輔助函數：解析公式字串
def _resolve_formula_string(formula_str, external_link_map):
    # === 新增：處理ArrayFormula對象和其他非字串類型 ===
    if hasattr(formula_str, 'text'):
        # ArrayFormula對象，提取text屬性
        formula_str = formula_str.text
    elif not isinstance(formula_str, str):
        # 其他非字串對象，嘗試轉換為字串
        try:
            formula_str = str(formula_str)
        except:
            return formula_str  # 無法轉換，返回原值
    
    # 確保是字串類型
    if not isinstance(formula_str, str):
        return formula_str
    
    for index_str, formatted_path in external_link_map.items():
        # 修復正則表達式：匹配 [1]WorksheetName! 的模式
        pattern = r'\[{}\]([^!]+)!'.format(re.escape(index_str))
        
        def replace_func(match):
            sheet_name = match.group(1)
            # 正確格式：'path\[file.xlsx]WorksheetName'!
            return f"{formatted_path}{sheet_name}'!"
        
        formula_str = re.sub(pattern, replace_func, formula_str)
    return formula_str


class ResolvedCellView:
    """
    包裝 openpyxl.Cell 物件，並在存取其值時解析外部連結。
    透過 __getattr__ 和 __setattr__ 代理所有未明確定義的屬性。
    """
    _wrapped_attrs = ('_cell', '_external_link_map') # 內部屬性列表

    def __init__(self, openpyxl_cell, external_link_map):
        object.__setattr__(self, '_cell', openpyxl_cell)
        object.__setattr__(self, '_external_link_map', external_link_map)

    @property
    def value(self):
        if self._cell.data_type == 'f':
            return _resolve_formula_string(self._cell.value, self._external_link_map)
        else:
            return self._cell.value

    @value.setter
    def value(self, new_value):
        self._cell.value = new_value

    # 明確定義常用屬性
    @property
    def coordinate(self):
        return self._cell.coordinate

    @property
    def row(self):
        return self._cell.row

    @property
    def column(self):
        return self._cell.column

    @property
    def data_type(self):
        return self._cell.data_type

    @property
    def font(self):
        return self._cell.font

    @property
    def fill(self):
        return self._cell.fill

    @property
    def border(self):
        return self._cell.border

    @property
    def alignment(self):
        return self._cell.alignment

    @property
    def number_format(self):
        return self._cell.number_format

    @number_format.setter
    def number_format(self, value):
        self._cell.number_format = value

    def __getattr__(self, name):
        # 代理所有未明確定義的屬性到底層 openpyxl.Cell
        return getattr(self._cell, name)

    def __setattr__(self, name, value):
        if name in self._wrapped_attrs:
            object.__setattr__(self, name, value)
        else:
            setattr(self._cell, name, value)


class ResolvedSheetView:
    """
    包裝 openpyxl.Worksheet 物件，並提供方法來獲取 ResolvedCellView 物件。
    透過 __getattr__ 和 __setattr__ 代理所有未明確定義的屬性。
    """
    _wrapped_attrs = ('_sheet', '_external_link_map') # 內部屬性列表

    def __init__(self, openpyxl_sheet, external_link_map):
        object.__setattr__(self, '_sheet', openpyxl_sheet)
        object.__setattr__(self, '_external_link_map', external_link_map)

    # 明確定義常用屬性
    @property
    def title(self):
        return self._sheet.title

    @property
    def min_row(self):
        return self._sheet.min_row

    @property
    def max_row(self):
        return self._sheet.max_row

    @property
    def min_column(self):
        return self._sheet.min_column

    @property
    def max_column(self):
        return self._sheet.max_column

    @property
    def column_dimensions(self):
        return self._sheet.column_dimensions

    @property
    def row_dimensions(self):
        return self._sheet.row_dimensions

    # 明確定義常用方法，並處理返回值的包裝
    def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None):
        for row in self._sheet.iter_rows(min_row, max_row, min_col, max_col):
            yield tuple(ResolvedCellView(cell, self._external_link_map) for cell in row)

    def __getitem__(self, key):
        cell = self._sheet[key]
        return ResolvedCellView(cell, self._external_link_map)

    def cell(self, row, column, value=None):
        original_cell = self._sheet.cell(row=row, column=column, value=value)
        return ResolvedCellView(original_cell, self._external_link_map)

    def append(self, iterable):
        self._sheet.append(iterable)

    def insert_rows(self, idx, amount=1):
        self._sheet.insert_rows(idx, amount)

    def delete_rows(self, idx, amount=1):
        self._sheet.delete_rows(idx, amount)

    def insert_cols(self, idx, amount=1):
        self._sheet.insert_cols(idx, amount)

    def delete_cols(self, idx, amount=1):
        self._sheet.delete_cols(idx, amount)

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        self._sheet.merge_cells(range_string, start_row, start_column, end_row, end_column)

    def unmerge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        self._sheet.unmerge_cells(range_string, start_row, start_column, end_row, end_column)

    def __getattr__(self, name):
        # 代理所有未明確定義的屬性到底層 openpyxl.Worksheet
        attr = getattr(self._sheet, name)
        if callable(attr):
            def wrapper(*args, **kwargs):
                result = attr(*args, **kwargs)
                # 如果方法返回 openpyxl.Cell，則包裝為 ResolvedCellView
                if isinstance(result, openpyxl.cell.cell.Cell):
                    return ResolvedCellView(result, self._external_link_map)
                return result
            return wrapper
        return attr

    def __setattr__(self, name, value):
        if name in self._wrapped_attrs:
            object.__setattr__(self, name, value)
        else:
            setattr(self._sheet, name, value)


class ResolvedWorkbookView:
    """
    包裝 openpyxl.Workbook 物件，並提供類似的介面，但其儲存格值會解析外部連結。
    透過 __getattr__ 和 __setattr__ 代理所有未明確定義的屬性。
    """
    _wrapped_attrs = ('_workbook', '_external_link_map') # 內部屬性列表

    def __init__(self, openpyxl_workbook):
        object.__setattr__(self, '_workbook', openpyxl_workbook)
        object.__setattr__(self, '_external_link_map', _get_external_link_map(openpyxl_workbook))

    # 明確定義常用屬性/方法，並處理返回值的包裝
    @property
    def active(self):
        return ResolvedSheetView(self._workbook.active, self._external_link_map)

    @property
    def sheetnames(self):
        return self._workbook.sheetnames

    def __getitem__(self, key):
        sheet = self._workbook[key]
        return ResolvedSheetView(sheet, self._external_link_map)

    def create_sheet(self, title=None, index=None):
        new_sheet = self._workbook.create_sheet(title=title, index=index)
        return ResolvedSheetView(new_sheet, self._external_link_map)

    def remove(self, worksheet):
        if isinstance(worksheet, ResolvedSheetView):
            self._workbook.remove(worksheet._sheet)
        else:
            self._workbook.remove(worksheet)

    def remove_sheet(self, worksheet):
        self.remove(worksheet)

    def get_sheet_by_name(self, name):
        sheet = self._workbook.get_sheet_by_name(name)
        if sheet:
            return ResolvedSheetView(sheet, self._external_link_map)
        return None

    def save(self, filename):
        self._workbook.save(filename)

    def __getattr__(self, name):
        # 代理所有未明確定義的屬性到底層 openpyxl.Workbook
        attr = getattr(self._workbook, name)
        if callable(attr):
            def wrapper(*args, **kwargs):
                result = attr(*args, **kwargs)
                # 如果方法返回 openpyxl.Worksheet，則包裝為 ResolvedSheetView
                if isinstance(result, openpyxl.worksheet.worksheet.Worksheet):
                    return ResolvedSheetView(result, self._external_link_map)
                return result
            return wrapper
        return attr

    def __setattr__(self, name, value):
        if name in self._wrapped_attrs:
            object.__setattr__(self, name, value)
        else:
            setattr(self._workbook, name, value)


def load_resolved_workbook(file_path, use_cache=True):
    """
    載入 Excel 檔案，並返回一個 ResolvedWorkbookView 物件。
    這個物件的儲存格值會自動解析外部連結，並代理所有未明確定義的屬性。
    
    Args:
        file_path: Excel 檔案路徑
        use_cache: 是否使用快取系統 (預設: True)
    """
    if use_cache:
        from .safe_cache import get_safe_cached_workbook
        workbook = get_safe_cached_workbook(file_path, data_only=False)
    else:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
    
    return ResolvedWorkbookView(workbook)


def read_cell_with_resolved_references(file_path, sheet_name, cell_address, use_cache=True):
    """
    使用 ResolvedWorkbookView 讀取指定 cell 的資訊
    返回: (formula, calculated_value, display_value, cell_type)
    """
    try:
        # 使用 resolved workbook 讀取 (with cache)
        resolved_wb = load_resolved_workbook(file_path, use_cache=use_cache)
        
        resolved_sheet = resolved_wb[sheet_name]
        resolved_cell = resolved_sheet[cell_address]
        
        # 獲取原始和解析後的值
        original_value = resolved_cell._cell.value
        resolved_value = resolved_cell.value
        cell_type = resolved_cell.data_type
        
        # 判斷 cell 類型和內容
        if cell_type == 'f':  # Formula
            formula = resolved_value  # 已經解析過 external references
            
            # 嘗試獲取計算值 (使用 data_only=True with cache)
            try:
                if use_cache:
                    from .safe_cache import get_safe_cached_workbook
                    data_wb = get_safe_cached_workbook(file_path, data_only=True)
                else:
                    data_wb = openpyxl.load_workbook(file_path, data_only=True)
                
                data_sheet = data_wb[sheet_name]
                data_cell = data_sheet[cell_address]
                calculated_value = data_cell.value
            except:
                calculated_value = "Cannot calculate"
            
            display_value = str(calculated_value) if calculated_value is not None else "N/A"
            
            # 檢查是否仍有未解析的 external references
            has_unresolved_refs = '[' in formula and ']' in formula and any(f'[{i}]' in formula for i in range(1, 10))
            
            return {
                'formula': formula,
                'calculated_value': calculated_value,
                'display_value': display_value,
                'cell_type': 'formula',
                'has_external_references': '[' in formula and ']' in formula
            }
        else:
            # 非公式 cell
            return {
                'formula': None,
                'calculated_value': resolved_value,
                'display_value': str(resolved_value) if resolved_value is not None else "",
                'cell_type': 'value',
                'has_external_references': False
            }
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'error': str(e),
            'formula': None,
            'calculated_value': None,
            'display_value': None,
            'cell_type': 'error',
            'has_external_references': False
        }