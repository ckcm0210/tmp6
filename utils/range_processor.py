# -*- coding: utf-8 -*-
"""
Range Processor - 處理Excel範圍地址，計算hash和維度信息
"""

import re
import hashlib
import openpyxl
from openpyxl.utils import range_boundaries
import os
import tempfile
import shutil

class RangeProcessor:
    """Excel範圍處理器"""
    
    def __init__(self):
        self.cache = {}  # 緩存已計算的hash
    
    def identify_ranges_in_formula(self, formula):
        """
        識別公式中的範圍地址
        
        Args:
            formula: Excel公式字符串
            
        Returns:
            list: 範圍信息列表
        """
        if not formula or not formula.startswith('='):
            return []
        
        ranges = []
        
        # 範圍模式：A1:B10, A:B, 1:5 等
        range_patterns = [
            r"([A-Z]+\d+):([A-Z]+\d+)",  # A1:B10 (儲存格範圍)
            r"([A-Z]+):([A-Z]+)",        # A:B (整列範圍)
            r"(\d+):(\d+)"               # 1:5 (整行範圍)
        ]
        
        for pattern in range_patterns:
            matches = re.findall(pattern, formula)
            for match in matches:
                start, end = match
                range_address = f"{start}:{end}"
                
                # 判斷範圍類型
                if re.match(r"[A-Z]+\d+", start) and re.match(r"[A-Z]+\d+", end):
                    range_type = "cell_range"
                elif re.match(r"[A-Z]+", start) and re.match(r"[A-Z]+", end):
                    range_type = "column_range"
                elif re.match(r"\d+", start) and re.match(r"\d+", end):
                    range_type = "row_range"
                else:
                    continue
                
                ranges.append({
                    'address': range_address,
                    'start': start,
                    'end': end,
                    'type': range_type
                })
        
        return ranges
    
    def calculate_range_dimensions(self, range_address):
        """
        計算範圍維度
        
        Args:
            range_address: 範圍地址 (如 A1:B10)
            
        Returns:
            dict: 維度信息
        """
        try:
            # 使用openpyxl解析範圍邊界
            min_col, min_row, max_col, max_row = range_boundaries(range_address)
            
            rows = max_row - min_row + 1
            columns = max_col - min_col + 1
            total_cells = rows * columns
            
            return {
                'rows': rows,
                'columns': columns,
                'total_cells': total_cells,
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col,
                'dimension_summary': f"{rows}行 x {columns}列"
            }
        except Exception as e:
            return {
                'rows': 0,
                'columns': 0,
                'total_cells': 0,
                'dimension_summary': f"無法解析範圍: {e}",
                'error': str(e)
            }
    
    def calculate_range_content_hash(self, workbook_path, sheet_name, range_address):
        """
        計算範圍內容的精確hash值
        
        Args:
            workbook_path: Excel文件路徑
            sheet_name: 工作表名稱
            range_address: 範圍地址
            
        Returns:
            dict: hash信息
        """
        cache_key = f"{workbook_path}|{sheet_name}|{range_address}"
        
        # 檢查緩存
        if cache_key in self.cache:
            return self.cache[cache_key]
        
        try:
            # 檢查文件是否存在
            if not os.path.exists(workbook_path):
                return {
                    'hash': 'FILE_NOT_FOUND',
                    'hash_short': 'FILE_NOT_FOUND',
                    'content_summary': '文件不存在',
                    'error': f'文件不存在: {workbook_path}'
                }
            
            # 使用原始檔案路徑（暫時還原，不使用臨時複本）
            wb = openpyxl.load_workbook(workbook_path, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                try:
                    wb.close()
                except:
                    pass
                return {
                    'hash': 'SHEET_NOT_FOUND',
                    'hash_short': 'SHEET_NOT_FOUND',
                    'content_summary': '工作表不存在',
                    'error': f'工作表不存在: {sheet_name}'
                }
            
            ws = wb[sheet_name]
            
            # 獲取範圍內的所有值
            range_cells = ws[range_address]
            
            # 收集所有值用於hash計算
            values = []
            value_types = {'number': 0, 'text': 0, 'formula': 0, 'empty': 0}
            
            # 處理單個儲存格的情況
            if not isinstance(range_cells, tuple):
                range_cells = ((range_cells,),)
            elif not isinstance(range_cells[0], tuple):
                range_cells = (range_cells,)
            
            for row in range_cells:
                for cell in row:
                    if cell.value is None:
                        values.append('')
                        value_types['empty'] += 1
                    elif isinstance(cell.value, (int, float)):
                        values.append(str(cell.value))
                        value_types['number'] += 1
                    elif isinstance(cell.value, str):
                        values.append(cell.value)
                        if cell.value.startswith('='):
                            value_types['formula'] += 1
                        else:
                            value_types['text'] += 1
                    else:
                        values.append(str(cell.value))
                        value_types['text'] += 1
            
            # 計算hash
            content_string = '|'.join(values)
            hash_object = hashlib.sha256(content_string.encode('utf-8'))
            full_hash = hash_object.hexdigest()
            short_hash = full_hash[:20]  # 前20位作為短hash，足夠做比較
            
            # 生成內容摘要
            total_cells = sum(value_types.values())
            non_empty = total_cells - value_types['empty']
            
            summary_parts = []
            if value_types['number'] > 0:
                summary_parts.append(f"{value_types['number']}數值")
            if value_types['text'] > 0:
                summary_parts.append(f"{value_types['text']}文字")
            if value_types['formula'] > 0:
                summary_parts.append(f"{value_types['formula']}公式")
            if value_types['empty'] > 0:
                summary_parts.append(f"{value_types['empty']}空白")
            
            content_summary = f"{non_empty}/{total_cells}非空 ({', '.join(summary_parts)})"
            
            result = {
                'hash': full_hash,
                'hash_short': short_hash,
                'content_summary': content_summary,
                'value_types': value_types,
                'total_values': len(values),
                'error': None
            }
            
            # 緩存結果
            self.cache[cache_key] = result
            
            # 關閉工作簿
            try:
                wb.close()
            except:
                pass
            return result
            
        except Exception as e:
            error_result = {
                'hash': 'ERROR',
                'hash_short': 'ERROR',
                'content_summary': f'讀取錯誤: {str(e)}',
                'error': str(e)
            }
            self.cache[cache_key] = error_result
            return error_result
    
    def process_range(self, workbook_path, sheet_name, range_address):
        """
        完整處理範圍：計算維度和hash
        
        Args:
            workbook_path: Excel文件路徑
            sheet_name: 工作表名稱  
            range_address: 範圍地址
            
        Returns:
            dict: 完整的範圍信息
        """
        # 計算維度
        dimensions = self.calculate_range_dimensions(range_address)
        
        # 計算hash
        hash_info = self.calculate_range_content_hash(workbook_path, sheet_name, range_address)
        
        # 合併信息
        result = {
            'address': range_address,
            'type': 'range',
            'workbook_path': workbook_path,
            'sheet_name': sheet_name,
            **dimensions,
            **hash_info
        }
        
        return result
    
    def clear_cache(self):
        """清除緩存"""
        self.cache.clear()


# 全局實例
range_processor = RangeProcessor()


def process_formula_ranges(formula, workbook_path, sheet_name):
    """
    便捷函數：處理公式中的所有範圍
    
    Args:
        formula: Excel公式
        workbook_path: Excel文件路徑
        sheet_name: 工作表名稱
        
    Returns:
        list: 處理後的範圍信息列表
    """
    ranges = range_processor.identify_ranges_in_formula(formula)
    processed_ranges = []
    
    for range_info in ranges:
        processed = range_processor.process_range(
            workbook_path, sheet_name, range_info['address']
        )
        processed_ranges.append(processed)
    
    return processed_ranges


# 測試函數
if __name__ == "__main__":
    # 測試範圍識別
    test_formula = "=SUM(A1:A100)+AVERAGE(B1:C10)"
    ranges = range_processor.identify_ranges_in_formula(test_formula)
    print("識別到的範圍:")
    for r in ranges:
        print(f"  {r}")
    
    # 測試維度計算
    dimensions = range_processor.calculate_range_dimensions("A1:C10")
    print(f"\nA1:C10 維度: {dimensions}")
    
    print("\nRange Processor ready for integration!")