import tkinter as tk
from tkinter import messagebox, ttk
import os
import re
import openpyxl
from core.excel_connector import activate_excel_window

def _perform_excel_selection(pane, affected_addresses):
    """
    Core logic for selecting ranges in Excel. This function is UI-agnostic.
    It directly interacts with the Excel application instance.
    """
    try:
        activate_excel_window(pane)
        pane.worksheet.Activate()
        if len(affected_addresses) == 1:
            pane.worksheet.Range(affected_addresses[0]).Select()
        else:
            # Create a union of all ranges for efficient selection
            range_objects = [pane.worksheet.Range(addr) for addr in affected_addresses]
            union_range = range_objects[0]
            for r in range_objects[1:]:
                union_range = pane.xl.Union(union_range, r)
            union_range.Select()
        return True, None  # Success
    except Exception as e:
        return False, str(e)  # Failure, return error message

def select_ranges_in_excel(summary_window, summary_tree, pane, link_to_addresses_cache):
    """
    UI-facing function that handles user interaction and calls the core selection logic.
    This function is responsible for all user dialogs (messagebox).
    """
    # --- UI Logic: Check for selected item ---
    selected_items = summary_tree.selection()
    if not selected_items:
        messagebox.showwarning("No Selection", "Please select an external link from the list first.", parent=summary_window)
        return

    selected_link = summary_tree.item(selected_items[0], "values")[0]
    
    # --- UI Logic: Check for affected addresses ---
    # This logic is now enhanced to handle both worksheet and workbook summary modes.
    affected_addresses = []
    # Check for an exact match first (worksheet mode)
    if selected_link in link_to_addresses_cache:
        affected_addresses.extend(link_to_addresses_cache[selected_link])
    else:
        # If no exact match, check for partial matches (workbook mode)
        for link_key, addresses in link_to_addresses_cache.items():
            if link_key.startswith(selected_link):
                affected_addresses.extend(addresses)

    if not affected_addresses:
        messagebox.showinfo("No Affected Cells", f"No cells were found for the selected link:\n{selected_link}", parent=summary_window)
        return

    # --- UI Logic: Check for Excel connection ---
    if not pane.xl or not pane.worksheet:
        messagebox.showerror("Excel Connection Error", "Not connected to Excel.", parent=summary_window)
        return

    # --- Call Core Logic ---
    success, error_message = _perform_excel_selection(pane, affected_addresses)

    # --- UI Logic: Handle result of core logic ---
    if not success:
        messagebox.showerror("Excel Operation Error", f"An error occurred while trying to select ranges in Excel:\n\n{error_message}", parent=summary_window)

def replace_links_in_excel(summary_window, replace_frame, pane, summary_tree, old_link_var, new_link_entry, rescan_var, formulas_to_summarize, link_to_addresses_cache, external_path_pattern, show_summary_by_workbook, show_summary_by_worksheet, current_mode, sorted_full_paths, btn_by_sheet, btn_by_workbook, browse_button, replace_button):
    calc_mode_prev = None
    calc_before_save_prev = None
    enable_events_prev = None
    old_link = old_link_var.get()
    newline = "\n"

    if old_link == "<No selection>":
        messagebox.showerror(
            "Replacement Failed - Old Link Not Selected",
            f"Reason: You have not selected an old link from the list.{newline}{newline}Please select an old link from the 'External Links' list and try again.",
            parent=summary_window
        )
        return

    new_link = new_link_entry.get().strip()
    if not new_link:
        messagebox.showerror(
            "Replacement Failed - New Link Is Empty",
            f"Reason: The 'New Link' input field cannot be empty.{newline}{newline}Please enter valid link information in the 'New Link' field.",
            parent=summary_window
        )
        return

    path_match = re.search(r"^(.*\\)?\[([^\]]+)\](.*)$", new_link)
    if not path_match:
        path_example1 = "C:\\path\\[filename.xlsx]Sheetname (with worksheet)"
        path_example2 = "C:\\path\\[filename.xlsx] (f[...)"
        messagebox.showerror(
            "Replacement Failed - New Link Format Error",
            f"The format of the 'New Link' is invalid and cannot be recognized.{newline}{newline}"
            f"Expected format examples:{newline} - {path_example1}{newline} - {path_example2}",
            parent=summary_window
        )
        return

    new_sheet_name_raw = path_match.groups()[2]
    new_sheet_name_cleaned = new_sheet_name_raw.strip("'")

    old_link_match = re.search(r"^(.*\\)?\[([^\]]+)\](.*)$", old_link)

    if not old_link_match:
        old_sheet_name_raw = ""
    else:
        old_sheet_name_raw = old_link_match.groups()[2]

    old_sheet_name_cleaned = old_sheet_name_raw.strip("'")

    if old_sheet_name_cleaned != new_sheet_name_cleaned:
        proceed = messagebox.askyesno(
            "Worksheet Name Mismatch",
            f"Warning: The worksheet name specified in your old link does not match that in the new link.{newline}{newline}"
            f"Old Link Worksheet: '{old_sheet_name_cleaned}'{newline}"
            f"New Link Worksheet: '{new_sheet_name_cleaned}'[...]",
            parent=summary_window
        )
        if not proceed:
            messagebox.showinfo("Operation Cancelled", "The replacement operation has been cancelled by the user.", parent=summary_window)
            return
    
    dir_path, file_name, sheet_name = path_match.groups()
    dir_path = dir_path if dir_path else os.path.dirname(pane.workbook.FullName)
    full_file_path = os.path.join(dir_path, file_name)

    if not old_sheet_name_cleaned and not new_sheet_name_cleaned:
        old_link_dir_path = ""
        old_link_file_name = ""

        if old_link_match:
            old_link_dir_path, old_link_file_name, _ = old_link_match.groups()
        
        old_full_file_path = os.path.join(old_link_dir_path if old_link_dir_path else os.path.dirname(pane.workbook.FullName), old_link_file_name)

        if not os.path.exists(old_full_file_path):
            messagebox.showerror(
                "Replacement Failed - Old Link File Not Found!",
                f"Reason: The file pointed to by the old link cannot be found.{newline}{newline}"
                f"Please check if the path is correct:{newline}'{old_full_file_path}'{newline}{newline}"
                f"Please ensure the old Excel file exists before per[...]",
                parent=summary_window
            )
            return
        if not os.path.exists(full_file_path):
            messagebox.showerror(
                "Replacement Failed - New Link File Not Found!",
                f"Reason: The file pointed to by the new link cannot be found.{newline}{newline}"
                f"Please check if the path is correct:{newline}'{full_file_path}'{newline}{newline}"
                f"Please ensure the new Excel file exists before perform[...]",
                parent=summary_window
            )
            return

        old_wb_sheetnames = set()
        new_wb_sheetnames = set()

        old_wb = None
        try:
            old_wb = openpyxl.load_workbook(old_full_file_path, read_only=True)
            old_wb_sheetnames = set(old_wb.sheetnames)
        except Exception as e:
            messagebox.showerror(
                "Replacement Failed - Unable to Read Old File!",
                f"Reason: An error occurred while trying to read the old link file. This may prevent correct validation of its worksheet names.{newline}{newline}"
                f"File Path: '{old_full_file_path}'{newline}Error Details:[...]",
                parent=summary_window
            )
            return
        finally:
            if old_wb:
                old_wb.close()
        
        new_wb = None
        try:
            new_wb = openpyxl.load_workbook(full_file_path, read_only=True)
            new_wb_sheetnames = set(new_wb.sheetnames)
        except Exception as e:
            messagebox.showerror(
                "Replacement Failed - Unable to Read New File!",
                f"Reason: An error occurred while trying to read the new link file. This may prevent correct validation of its worksheet names.{newline}{newline}"
                f"File Path: '{full_file_path}'{newline}Error Details: {e}[...]",
                parent=summary_window
            )
            return
        finally:
            if new_wb:
                new_wb.close()

        # Check if new file contains all worksheets that are actually used in the selected external links
        used_worksheets = set()

        # Find all formulas that are affected by this replacement
        affected_formulas_content = []
        try:
            formula_idx = pane.view.tree_columns.index("formula")
        except ValueError:
            formula_idx = 2 # fallback
        
        for formula_data in formulas_to_summarize:
            if len(formula_data) > formula_idx:
                formula_content = str(formula_data[formula_idx])
                # We only care about formulas related to the specific old_link being replaced
                if old_link in formula_content:
                    affected_formulas_content.append(formula_content)

        # From these affected formulas, find all unique worksheet names they refer to.
        # A single formula might have multiple different external links.
        for formula_content in affected_formulas_content:
            matches = external_path_pattern.findall(formula_content)
            for match in matches:
                if ']' in match:
                    try:
                        worksheet_part = match.split(']', 1)[1].strip("'")
                        if worksheet_part:
                            used_worksheets.add(worksheet_part)
                    except IndexError:
                        pass
        
        # Now check if the new workbook has all the worksheets required by the affected formulas
        if used_worksheets: # Only check if there are any worksheets to verify
            missing_worksheets = used_worksheets - new_wb_sheetnames
            if missing_worksheets:
                missing_list = ", ".join(sorted(list(missing_worksheets)))
                new_list = ", ".join(sorted(list(new_wb_sheetnames))) if new_wb_sheetnames else "None"
                messagebox.showerror(
                    "Replacement Failed - Required Worksheets Not Found!",
                    f"The new file does not contain all worksheets required by the formulas you are trying to update.{newline}{newline}"
                    f"Missing Worksheets in new file: {missing_list}{newline}{newline}"
                    f"Worksheets available in new file: {new_list}",
                    parent=summary_window
                )
                return
    
    if not os.path.exists(full_file_path):
        messagebox.showerror(
            "Replacement Failed - New Link File Not Found!",
            f"Reason: The file pointed to by the new link cannot be found.{newline}{newline}"
            f"Please check if the path is correct:{newline}'{full_file_path}'{newline}{newline}"
            f"Please ensure the new Excel file exists before performing [...]",
            parent=summary_window
        )
        return

    try:
        wb = openpyxl.load_workbook(full_file_path, read_only=True)
        cleaned_sheet_name = sheet_name.strip("'")
        if sheet_name and cleaned_sheet_name not in wb.sheetnames:
            messagebox.showerror(
                "Replacement Failed - New File Worksheet Not Found!",
                f'Reason: The worksheet "{cleaned_sheet_name}" specified in the new link was not found in the target file "{file_name}".{newline}{newline}Please check if the worksheet name in [...]',
                parent=summary_window
            )
            wb.close()
            return
        wb.close()
    except Exception as e:
        messagebox.showerror(
            "Replacement Failed - Unable to Read New File!",
            f"Reason: An error occurred while trying to read the new link file.{newline}{newline}File Path: '{full_file_path}'{newline}Error Details: {e}{newline}{newline}Please ensure the fil[...]",
            parent=summary_window
        )
        return

    if not pane.worksheet:
        messagebox.showerror(
            "Replacement Failed - Not Connected to Excel!",
            f"Reason: The tool is not successfully connected to a live Excel worksheet, thus unable to perform update operations.{newline}{newline}Please ensure you have an Excel file open and[...]",
            parent=summary_window
        )
        return
    
    affected_cells = []
    formula_idx = pane.view.tree_columns.index("formula")
    address_idx = pane.view.tree_columns.index("address")

    current_formulas = [pane.view.result_tree.item(item_id, "values") for item_id in pane.view.result_tree.get_children()]

    for item_data in current_formulas:
        if len(item_data) > formula_idx and old_link in str(item_data[formula_idx]):
            address = item_data[address_idx]
            formula = item_data[formula_idx]
            affected_cells.append((address, formula))

    if not affected_cells:
        messagebox.showinfo("No Link Found", "The selected old link was not found in any formula in the current view.", parent=summary_window)
        return

    try:
        worksheet_name_snapshot = pane.worksheet.Name
    except Exception:
        worksheet_name_snapshot = "<Unknown Worksheet>"
    
    confirmation = messagebox.askyesno(
        "Confirm Replacement Operation",
        f"You are about to replace the following link:{newline}{newline}Old Link: {old_link}{newline}{newline}New Link: {new_link}{newline}{newline}This will affect {len(affected_cells)} cells[...]",
        parent=summary_window
    )

    if not confirmation:
        messagebox.showinfo("Operation Cancelled", "The replacement operation has been cancelled by the user.", parent=summary_window)
        return

    # Lock UI during replacement
    replace_button.configure(state='disabled')
    btn_by_sheet.configure(state='disabled')
    btn_by_workbook.configure(state='disabled')
    browse_button.configure(state='disabled')
    summary_window.configure(cursor='wait')
    
    # Create progress frame
    progress_frame = ttk.Frame(replace_frame)
    progress_frame.grid(row=3, column=0, columnspan=3, sticky="ew", padx=5, pady=5)
    progress_frame.columnconfigure(0, weight=1)
    
    progress_label = ttk.Label(progress_frame, text="Preparing replacement...")
    progress_label.grid(row=0, column=0, sticky="w")
    
    progress_bar = ttk.Progressbar(progress_frame, mode='determinate', length=300)
    progress_bar.grid(row=1, column=0, sticky="ew", pady=(2, 0))
    progress_bar['maximum'] = 100
    
    summary_window.update_idletasks()

    # Start: Apply Excel application settings for performance
    if pane.xl:
        try:
            calc_mode_prev = pane.xl.Calculation
            calc_before_save_prev = pane.xl.Application.CalculateBeforeSave
            enable_events_prev = pane.xl.Application.EnableEvents
            interactive_prev = pane.xl.Application.Interactive
            
            pane.xl.Calculation = -4135  # xlCalculationManual
            pane.xl.Application.CalculateBeforeSave = False
            pane.xl.Application.EnableEvents = False
            pane.xl.Application.Interactive = False  # Prevent user interaction with Excel
        except Exception as e:
            print(f"Failed to set Excel application properties: {e}")
            calc_mode_prev = None
            calc_before_save_prev = None
            enable_events_prev = None
            interactive_prev = None
    # End: Apply Excel application settings for performance

    activate_excel_window(pane)
    pane.worksheet.Activate()

    # Batch processing with progress updates
    total_updated_count = 0
    total_error_count = 0
    total_cells = len(affected_cells)
    batch_size = 50

    for i in range(0, total_cells, batch_size):
        batch = affected_cells[i:i+batch_size]
        batch_end = min(i + batch_size, total_cells)

        # Update progress
        progress_label.config(text=f"Processing cells {i+1}-{batch_end} of {total_cells}...")
        progress_bar['value'] = (i / total_cells) * 100
        summary_window.update_idletasks()

        # Prepare updates for the core function
        updates_for_batch = []
        for address, old_formula in batch:
            new_formula = old_formula.replace(old_link, new_link)
            updates_for_batch.append((address, new_formula))

        # Call the core replacement logic for the batch
        updated_count, error_count = _perform_excel_formula_updates(pane, updates_for_batch)
        total_updated_count += updated_count
        total_error_count += error_count

    # Final progress update
    progress_label.config(text="Replacement completed!")
    progress_bar['value'] = 100
    summary_window.update_idletasks()

    # Restore Excel application settings
    if pane.xl and calc_mode_prev is not None:
        try:
            pane.xl.Application.Interactive = True if interactive_prev is None else interactive_prev
            pane.xl.Application.EnableEvents = True if enable_events_prev is None else enable_events_prev
            pane.xl.Application.CalculateBeforeSave = True if calc_before_save_prev is None else calc_before_save_prev
            pane.xl.Calculation = -4105 if calc_mode_prev is None else calc_mode_prev # xlCalculationAutomatic
            pane.xl.CalculateFullRebuild()
        except Exception as e:
            print(f"Failed to restore Excel application properties: {e}")

    # Unlock UI
    replace_button.configure(state='normal')
    btn_by_sheet.configure(state='normal')
    btn_by_workbook.configure(state='normal')
    browse_button.configure(state='normal')
    summary_window.configure(cursor='')
    
    # Remove progress frame
    progress_frame.destroy()

    messagebox.showinfo(
        "Replacement Complete",
        f"Link replacement operation has finished.{newline}{newline}Successfully updated: {total_updated_count} cells{newline}Failed to update: {total_error_count} cells{newline}{newline}Please re-scan the worksheet to see the updated links.",
        parent=summary_window
    )

    # Re-scan the current summarize range from actual Excel formulas
    unique_full_paths = set()
    for formula_data in formulas_to_summarize:
        if len(formula_data) > address_idx:
            cell_address = formula_data[address_idx]
            try:
                # Read the actual formula from Excel after replacement
                actual_formula = pane.worksheet.Range(cell_address).Formula
                matches = external_path_pattern.findall(str(actual_formula))
                if matches:
                    unique_full_paths.update(matches)
            except Exception:
                # If we can't read from Excel, fall back to tree data
                if len(formula_data) > formula_idx:
                    formula_content = formula_data[formula_idx]
                    matches = external_path_pattern.findall(str(formula_content))
                    if matches:
                        unique_full_paths.update(matches)
    sorted_full_paths[:] = sorted(list(unique_full_paths))
    # Stay in current mode after replacement
    if current_mode == "workbook":
        show_summary_by_workbook()
    else:
        show_summary_by_worksheet()

    old_link_var.set("<No selection>")
    new_link_entry.delete(0, 'end')
    summary_window.did_replace = True

def _perform_excel_formula_updates(pane, updates):
    """
    Applies a list of formula updates to Excel.
    'updates' is a list of (address, new_formula) tuples.
    Returns (updated_count, error_count).
    """
    updated_count = 0
    error_count = 0
    for address, new_formula in updates:
        try:
            pane.worksheet.Range(address).Formula = new_formula
            updated_count += 1
        except Exception:
            error_count += 1
    return updated_count, error_count